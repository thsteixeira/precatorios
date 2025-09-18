from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, StreamingHttpResponse
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Sum
from django.utils import timezone
from django.core.management import call_command
from django.urls import reverse
from django.conf import settings
import tempfile
import os
import logging
import mimetypes
from io import StringIO
from .models import Precatorio, Cliente, Alvara, Requerimento, Fase, Tipo, FaseHonorariosContratuais, FaseHonorariosSucumbenciais, TipoDiligencia, Diligencias, PedidoRequerimento, ContaBancaria, Recebimentos
from .forms import (
    PrecatorioForm, ClienteForm, PrecatorioSearchForm, 
    ClienteSearchForm, RequerimentoForm, ClienteSimpleForm, 
    AlvaraSimpleForm, FaseForm, TipoForm, FaseHonorariosContratuaisForm, FaseHonorariosSucumbenciaisForm, TipoDiligenciaForm,
    DiligenciasForm, DiligenciasUpdateForm, PedidoRequerimentoForm, ContaBancariaForm, RecebimentosForm
)

logger = logging.getLogger(__name__)

# ===============================
# UTILITY FUNCTIONS
# ===============================

def get_acordo_pedido_names():
    """Get dynamic list of acordo pedido names from database"""
    return list(
        PedidoRequerimento.objects.filter(
            nome__icontains='Acordo',
            ativo=True
        ).values_list('nome', flat=True)
    )

def get_prioridade_pedido_names():
    """Get dynamic list of prioridade pedido names from database"""
    return list(
        PedidoRequerimento.objects.filter(
            nome__icontains='Prioridade',
            ativo=True
        ).values_list('nome', flat=True)
    )

# ===============================
# AUTHENTICATION VIEWS
# ===============================

def login_view(request):
    """Custom login view"""
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bem-vindo, {user.get_full_name() or user.username}!')
                # Redirect to next URL if provided, otherwise to home
                next_url = request.GET.get('next', 'home')
                return redirect(next_url)
            else:
                messages.error(request, 'Credenciais inválidas.')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = AuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})

def logout_view(request):
    """Custom logout view"""
    logout(request)
    messages.info(request, 'Você foi desconectado com sucesso.')
    return redirect('login')

# ===============================
# HOME VIEWS
# ===============================

@login_required
def home_view(request):
    """Home view with dashboard statistics"""
    from django.db.models import Sum, Count
    
    # Get counts for dashboard
    total_precatorios = Precatorio.objects.count()
    total_clientes = Cliente.objects.count()
    total_alvaras = Alvara.objects.count()
    total_requerimentos = Requerimento.objects.count()
    total_diligencias = Diligencias.objects.count()
    total_tipos_diligencia = TipoDiligencia.objects.filter(ativo=True).count()
    
    # Get financial statistics - using correct field names
    total_valor_precatorios = Precatorio.objects.aggregate(total=Sum('valor_de_face'))['total'] or 0
    total_valor_alvaras = Alvara.objects.aggregate(
        principal=Sum('valor_principal'),
        contratuais=Sum('honorarios_contratuais'),
        sucumbenciais=Sum('honorarios_sucumbenciais')
    )
    
    valor_alvaras = (
        (total_valor_alvaras['principal'] or 0) + 
        (total_valor_alvaras['contratuais'] or 0) + 
        (total_valor_alvaras['sucumbenciais'] or 0)
    )
    
    total_valor_requerimentos = Requerimento.objects.aggregate(total=Sum('valor'))['total'] or 0
    
    # Get diligencias statistics
    diligencias_pendentes = Diligencias.objects.filter(concluida=False).count()
    diligencias_concluidas = Diligencias.objects.filter(concluida=True).count()
    diligencias_atrasadas = Diligencias.objects.filter(
        concluida=False,
        data_final__lt=timezone.now().date()
    ).count()
    diligencias_urgentes = Diligencias.objects.filter(
        concluida=False,
        urgencia='alta'
    ).count()
    
    # Get recent activity
    recent_precatorios = Precatorio.objects.prefetch_related('clientes').order_by('cnj')[:5]
    recent_alvaras = Alvara.objects.select_related('cliente', 'precatorio').order_by('-id')[:5]
    recent_requerimentos = Requerimento.objects.select_related('cliente', 'precatorio').order_by('-id')[:5]
    recent_diligencias = Diligencias.objects.select_related('cliente', 'tipo', 'responsavel').order_by('-data_criacao')[:5]
    
    context = {
        'total_precatorios': total_precatorios,
        'total_clientes': total_clientes,
        'total_alvaras': total_alvaras,
        'total_requerimentos': total_requerimentos,
        'total_diligencias': total_diligencias,
        'total_tipos_diligencia': total_tipos_diligencia,
        'total_valor_precatorios': total_valor_precatorios,
        'valor_alvaras': valor_alvaras,
        'total_valor_requerimentos': total_valor_requerimentos,
        'diligencias_pendentes': diligencias_pendentes,
        'diligencias_concluidas': diligencias_concluidas,
        'diligencias_atrasadas': diligencias_atrasadas,
        'diligencias_urgentes': diligencias_urgentes,
        'recent_precatorios': recent_precatorios,
        'recent_alvaras': recent_alvaras,
        'recent_requerimentos': recent_requerimentos,
        'recent_diligencias': recent_diligencias,
    }
    
    return render(request, 'precapp/home.html', context)


# ===============================
# PRECATORIO VIEWS
# ===============================

@login_required
def novoPrec_view(request):
    """View to create a new Precatorio"""
    if request.method == 'POST':
        form = PrecatorioForm(request.POST, request.FILES)
        if form.is_valid():
            precatorio = form.save()
            
            # Store the original filename if an integra file was uploaded
            if 'integra_precatorio' in request.FILES:
                uploaded_file = request.FILES['integra_precatorio']
                original_filename = uploaded_file.name
                precatorio.integra_precatorio_filename = original_filename
                precatorio.save()
                logger.info(f"Stored original filename for new precatorio {precatorio.cnj}: {original_filename}")
            
            messages.success(request, f'Precatório {precatorio.cnj} criado com sucesso!')
            return redirect('precatorios')  # Redirect to the list view after successful creation
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = PrecatorioForm()
    
    return render(request, 'precapp/novo_precatorio.html', {'form': form})
    

@login_required
def precatorio_view(request):
    """View to display all precatorios with filtering support"""
    precatorios = Precatorio.objects.all().prefetch_related('requerimento_set', 'requerimento_set__fase')
    
    # Apply filters based on GET parameters
    cnj_filter = request.GET.get('cnj', '').strip()
    origem_filter = request.GET.get('origem', '').strip()
    orcamento_filter = request.GET.get('orcamento', '').strip()
    credito_principal_filter = request.GET.get('credito_principal', '')
    honorarios_contratuais_filter = request.GET.get('honorarios_contratuais', '')
    honorarios_sucumbenciais_filter = request.GET.get('honorarios_sucumbenciais', '')
    tipo_filter = request.GET.get('tipo', '')
    requerimento_filter = request.GET.get('requerimento', '')
    status_requerimento_filter = request.GET.get('status_requerimento', '')
    
    if cnj_filter:
        precatorios = precatorios.filter(cnj__icontains=cnj_filter)
    
    if origem_filter:
        precatorios = precatorios.filter(origem__icontains=origem_filter)
    
    if orcamento_filter:
        try:
            orcamento_year = int(orcamento_filter)
            precatorios = precatorios.filter(orcamento=orcamento_year)
        except ValueError:
            # If orcamento_filter is not a valid integer, ignore the filter
            pass
    
    if credito_principal_filter:
        precatorios = precatorios.filter(credito_principal=credito_principal_filter)
    
    if honorarios_contratuais_filter:
        precatorios = precatorios.filter(honorarios_contratuais=honorarios_contratuais_filter)
    
    if honorarios_sucumbenciais_filter:
        precatorios = precatorios.filter(honorarios_sucumbenciais=honorarios_sucumbenciais_filter)
    
    if tipo_filter:
        precatorios = precatorios.filter(tipo_id=tipo_filter)
    
    # Apply new dynamic requerimento filters
    if requerimento_filter:
        if requerimento_filter == 'sem_requerimento':
            # Filter for precatorios that have NO requerimentos at all
            precatorios_com_requerimento = Requerimento.objects.values_list('precatorio__cnj', flat=True).distinct()
            precatorios = precatorios.exclude(cnj__in=precatorios_com_requerimento)
        else:
            # Filter by specific requerimento type (pedido)
            precatorios_com_pedido = Requerimento.objects.filter(
                pedido_id=requerimento_filter
            ).values_list('precatorio__cnj', flat=True).distinct()
            precatorios = precatorios.filter(cnj__in=precatorios_com_pedido)
    
    if status_requerimento_filter:
        # Filter by requerimento status (fase)
        precatorios_com_fase = Requerimento.objects.filter(
            fase_id=status_requerimento_filter
        ).values_list('precatorio__cnj', flat=True).distinct()
        precatorios = precatorios.filter(cnj__in=precatorios_com_fase)
    
    # Calculate summary statistics
    total_precatorios = precatorios.count()
    
    # Calculate counts for different payment statuses
    pendentes_principal = precatorios.filter(credito_principal='pendente').count()
    quitados_principal = precatorios.filter(credito_principal='quitado').count()
    parciais_principal = precatorios.filter(credito_principal='parcial').count()
    vendidos_principal = precatorios.filter(credito_principal='vendido').count()
    
    # Calculate prioritarios count (precatorios with prioridade requerimentos)
    # Get all pedidos that contain "Prioridade" in the name
    prioridade_pedidos = PedidoRequerimento.objects.filter(
        nome__icontains='Prioridade', 
        ativo=True
    ).values_list('id', flat=True)
    
    prioritarios_cnjs = Requerimento.objects.filter(
        pedido_id__in=prioridade_pedidos
    ).values_list('precatorio__cnj', flat=True).distinct()
    prioritarios = precatorios.filter(cnj__in=prioritarios_cnjs).count()
    
    # Calculate total value of displayed precatorios
    total_valor_precatorios = sum(precatorio.valor_de_face or 0 for precatorio in precatorios)
    
    # Get all active tipos for the filter dropdown
    tipos = Tipo.get_tipos_ativos()
    
    # Get all active pedidos and fases for dynamic dropdowns
    all_pedidos = PedidoRequerimento.objects.filter(ativo=True).order_by('nome')
    all_fases = Fase.objects.filter(
        ativa=True, 
        tipo__in=['requerimento', 'ambos']
    ).order_by('nome')
    
    context = {
        'precatorios': precatorios,
        'total_precatorios': total_precatorios,
        'total_valor_precatorios': total_valor_precatorios,
        'pendentes_principal': pendentes_principal,
        'quitados_principal': quitados_principal,
        'parciais_principal': parciais_principal,
        'vendidos_principal': vendidos_principal,
        'prioritarios': prioritarios,
        'tipos': tipos,
        'all_pedidos': all_pedidos,
        'all_fases': all_fases,
        'status_choices': Precatorio.STATUS_PAGAMENTO_CHOICES,  # Add status choices for template
        # Include current filter values to maintain state in form
        'current_cnj': cnj_filter,
        'current_origem': origem_filter,
        'current_orcamento': orcamento_filter,
        'current_credito_principal': credito_principal_filter,
        'current_honorarios_contratuais': honorarios_contratuais_filter,
        'current_honorarios_sucumbenciais': honorarios_sucumbenciais_filter,
        'current_tipo': tipo_filter,
        'current_requerimento': requerimento_filter,
        'current_status_requerimento': status_requerimento_filter,
    }
    
    return render(request, 'precapp/precatorio_list.html', context)


@login_required
def precatorio_detalhe_view(request, precatorio_cnj):
    """View to display and edit a single precatorio and manage client relationships"""
    precatorio = get_object_or_404(Precatorio, cnj=precatorio_cnj)
    
    # Get all clients associated with this precatorio via many-to-many relationship
    associated_clientes = precatorio.clientes.all().order_by('nome')
    
    # Get all alvarás associated with this precatorio
    alvaras = Alvara.objects.filter(precatorio=precatorio).select_related(
        'cliente', 'fase', 'fase_honorarios_contratuais'
    ).order_by('-id')
    
    # Calculate alvarás totals
    alvaras_total_principal = sum(alvara.valor_principal or 0 for alvara in alvaras)
    alvaras_total_contratuais = sum(alvara.honorarios_contratuais or 0 for alvara in alvaras)
    alvaras_total_sucumbenciais = sum(alvara.honorarios_sucumbenciais or 0 for alvara in alvaras)

    # Get all requerimentos associated with this precatorio
    requerimentos = Requerimento.objects.filter(precatorio=precatorio).select_related(
        'cliente', 'fase', 'pedido'
    ).order_by('-id')
    
    # Initialize forms
    precatorio_form = None
    client_search_form = ClienteSearchForm()
    requerimento_form = RequerimentoForm(precatorio=precatorio)
    cliente_form = ClienteSimpleForm()
    alvara_form = AlvaraSimpleForm(precatorio=precatorio)
    
    if request.method == 'POST':
        if 'edit_precatorio' in request.POST:
            # Handle precatorio editing with explicit file management
            old_file = None
            if precatorio.integra_precatorio:
                old_file = precatorio.integra_precatorio.name
                
            precatorio_form = PrecatorioForm(request.POST, request.FILES, instance=precatorio)
            if precatorio_form.is_valid():
                # Check if a new file was uploaded
                new_file_uploaded = 'integra_precatorio' in request.FILES
                
                if new_file_uploaded and old_file:
                    # Delete the old file before saving the new one
                    try:
                        from django.core.files.storage import default_storage
                        if default_storage.exists(old_file):
                            default_storage.delete(old_file)
                            logger.info(f"Manually deleted old file: {old_file}")
                    except Exception as e:
                        logger.error(f"Error deleting old file manually: {str(e)}")
                
                # Save the form
                updated_precatorio = precatorio_form.save(commit=False)
                
                # Store the original filename if a new file was uploaded
                if new_file_uploaded:
                    uploaded_file = request.FILES['integra_precatorio']
                    updated_precatorio.integra_precatorio_filename = uploaded_file.name
                    logger.info(f"Stored original filename: {uploaded_file.name}")
                
                # Save the instance
                updated_precatorio.save()
                
                # Force refresh the instance to ensure we have the latest file info
                updated_precatorio.refresh_from_db()
                
                if new_file_uploaded:
                    messages.success(request, f'Precatório {precatorio.cnj} atualizado com sucesso! Novo arquivo "{uploaded_file.name}" carregado.')
                else:
                    messages.success(request, f'Precatório {precatorio.cnj} atualizado com sucesso!')
                    
                return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            else:
                messages.error(request, 'Por favor, corrija os erros abaixo.')
        
        elif 'update_payment_status' in request.POST:
            # Handle inline payment status update
            credito_principal = request.POST.get('credito_principal', '').strip()
            honorarios_contratuais = request.POST.get('honorarios_contratuais', '').strip()
            honorarios_sucumbenciais = request.POST.get('honorarios_sucumbenciais', '').strip()
            
            # Update the payment status fields
            if credito_principal:
                precatorio.credito_principal = credito_principal
            if honorarios_contratuais:
                precatorio.honorarios_contratuais = honorarios_contratuais
            if honorarios_sucumbenciais:
                precatorio.honorarios_sucumbenciais = honorarios_sucumbenciais
            
            precatorio.save(update_fields=['credito_principal', 'honorarios_contratuais', 'honorarios_sucumbenciais'])
            messages.success(request, 'Status dos Pagamentos atualizado com sucesso!')
            return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            
        elif 'update_observacao' in request.POST:
            # Handle inline observacao update
            observacao = request.POST.get('observacao', '').strip()
            precatorio.observacao = observacao
            precatorio.save(update_fields=['observacao'])
            if observacao:
                messages.success(request, 'Observações atualizadas com sucesso!')
            else:
                messages.success(request, 'Observações removidas com sucesso!')
            return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            
        elif 'update_file' in request.POST:
            # Handle inline file update
            if 'integra_precatorio' in request.FILES:
                old_file = None
                if precatorio.integra_precatorio:
                    old_file = precatorio.integra_precatorio.name
                
                # Delete the old file before saving the new one
                if old_file:
                    try:
                        from django.core.files.storage import default_storage
                        if default_storage.exists(old_file):
                            default_storage.delete(old_file)
                            logger.info(f"Manually deleted old file: {old_file}")
                    except Exception as e:
                        logger.error(f"Error deleting old file manually: {str(e)}")
                
                # Save the new file
                uploaded_file = request.FILES['integra_precatorio']
                precatorio.integra_precatorio = uploaded_file
                precatorio.integra_precatorio_filename = uploaded_file.name
                precatorio.save(update_fields=['integra_precatorio', 'integra_precatorio_filename'])
                
                messages.success(request, f'Arquivo "{uploaded_file.name}" enviado com sucesso!')
                logger.info(f"Stored original filename: {uploaded_file.name}")
            else:
                messages.error(request, 'Nenhum arquivo foi selecionado.')
            return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            
        elif 'delete_file' in request.POST:
            # Handle inline file deletion
            if precatorio.integra_precatorio:
                old_file = precatorio.integra_precatorio.name
                filename = precatorio.integra_precatorio_filename or "arquivo"
                
                # Delete the file from storage
                try:
                    from django.core.files.storage import default_storage
                    if default_storage.exists(old_file):
                        default_storage.delete(old_file)
                        logger.info(f"Deleted file: {old_file}")
                except Exception as e:
                    logger.error(f"Error deleting file: {str(e)}")
                
                # Clear the file fields
                precatorio.integra_precatorio = None
                precatorio.integra_precatorio_filename = None
                precatorio.save(update_fields=['integra_precatorio', 'integra_precatorio_filename'])
                
                messages.success(request, f'Arquivo "{filename}" excluído com sucesso!')
            else:
                messages.warning(request, 'Nenhum arquivo para excluir.')
            return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
        
        elif 'link_cliente' in request.POST:
            # Handle client linking
            client_search_form = ClienteSearchForm(request.POST)  # This creates the form with POST data
            if client_search_form.is_valid():
                cpf = client_search_form.cleaned_data['cpf']
                try:
                    cliente = Cliente.objects.get(cpf=cpf)
                    if cliente in associated_clientes:
                        messages.warning(request, f'O cliente {cliente.nome} já está vinculado ao precatório {precatorio.cnj}.')
                    else:
                        # Add the client to the precatorio
                        precatorio.clientes.add(cliente)
                        messages.success(request, f'Cliente {cliente.nome} vinculado com sucesso ao precatório {precatorio.cnj}!')
                        return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                except Cliente.DoesNotExist:
                    messages.error(request, f'Cliente com CPF {cpf} não encontrado. Verifique se o CPF está correto e se o cliente está cadastrado.')
            else:
                messages.error(request, 'Por favor, corrija os erros no CPF.')
            # Note: client_search_form now contains the POST data and any validation errors
        
        elif 'unlink_cliente' in request.POST:
            # Handle client unlinking
            cliente_cpf = request.POST.get('cliente_cpf')
            try:
                cliente = Cliente.objects.get(cpf=cliente_cpf)
                if cliente in associated_clientes:
                    precatorio.clientes.remove(cliente)
                    messages.success(request, f'Cliente {cliente.nome} desvinculado do precatório {precatorio.cnj}.')
                    return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                else:
                    messages.error(request, 'Este cliente não está vinculado a este precatório.')
            except Cliente.DoesNotExist:
                                    messages.error(request, 'Cliente não encontrado.')
                    
        elif 'create_requerimento' in request.POST:
            # Handle requerimento creation
            requerimento_form = RequerimentoForm(request.POST, precatorio=precatorio)
            if requerimento_form.is_valid():
                cpf = requerimento_form.cleaned_data['cliente_cpf']
                # Remove dots and dashes, keep only numbers
                cpf_numbers = ''.join(filter(str.isdigit, cpf))
                try:
                    cliente = Cliente.objects.get(cpf=cpf_numbers)
                    # Create the requerimento
                    requerimento = requerimento_form.save(commit=False)
                    requerimento.precatorio = precatorio
                    requerimento.cliente = cliente
                    requerimento.save()
                    messages.success(request, f'Requerimento criado com sucesso para {cliente.nome}!')
                    return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                except Cliente.DoesNotExist:
                    messages.error(request, f'Cliente com CPF {cpf} não encontrado. Verifique se o CPF está correto e se o cliente está cadastrado.')
            else:
                messages.error(request, 'Por favor, corrija os erros no formulário do requerimento.')
                
        elif 'create_cliente' in request.POST:
            # Handle client creation
            cliente_form = ClienteSimpleForm(request.POST)
            if cliente_form.is_valid():
                cliente = cliente_form.save()
                # Automatically link the new client to this precatorio
                precatorio.clientes.add(cliente)
                messages.success(request, f'Cliente {cliente.nome} criado e vinculado com sucesso ao precatório!')
                return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            else:
                messages.error(request, 'Por favor, corrija os erros no formulário do cliente.')
                
        elif 'create_alvara' in request.POST:
            # Handle alvara creation
            alvara_form = AlvaraSimpleForm(request.POST, precatorio=precatorio)
            if alvara_form.is_valid():
                cpf = alvara_form.cleaned_data['cliente_cpf']
                # Remove dots and dashes, keep only numbers
                cpf_numbers = ''.join(filter(str.isdigit, cpf))
                try:
                    cliente = Cliente.objects.get(cpf=cpf_numbers)
                    # Create the alvara
                    alvara = alvara_form.save(commit=False)
                    alvara.precatorio = precatorio
                    alvara.cliente = cliente
                    alvara.save()
                    messages.success(request, f'Alvará criado com sucesso para {cliente.nome}!')
                    # When creating a new alvará, always show the fase changed modal
                    return redirect(f"{reverse('precatorio_detalhe', kwargs={'precatorio_cnj': precatorio.cnj})}?fase_changed=true")
                except Cliente.DoesNotExist:
                    messages.error(request, f'Cliente com CPF {cpf} não encontrado. Verifique se o CPF está correto e se o cliente está cadastrado.')
            else:
                messages.error(request, 'Por favor, corrija os erros no formulário do alvará.')
        
        elif 'update_alvara' in request.POST:
            # Handle alvara update
            alvara_id = request.POST.get('alvara_id')
            try:
                alvara = get_object_or_404(Alvara, id=alvara_id, precatorio=precatorio)
                
                # Track original fase values to detect changes
                original_fase = alvara.fase
                original_fase_honorarios_contratuais = alvara.fase_honorarios_contratuais
                original_fase_honorarios_sucumbenciais = alvara.fase_honorarios_sucumbenciais
                
                # Update the alvara fields
                alvara.valor_principal = float(request.POST.get('valor_principal', alvara.valor_principal))
                alvara.honorarios_contratuais = float(request.POST.get('honorarios_contratuais', alvara.honorarios_contratuais))
                alvara.honorarios_sucumbenciais = float(request.POST.get('honorarios_sucumbenciais', alvara.honorarios_sucumbenciais))
                alvara.tipo = request.POST.get('tipo', alvara.tipo)
                
                # Handle fase field properly (ForeignKey)
                fase_id = request.POST.get('fase')
                if fase_id:
                    try:
                        fase = Fase.objects.get(id=fase_id)
                        alvara.fase = fase
                    except Fase.DoesNotExist:
                        messages.error(request, 'Fase selecionada não existe.')
                        return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                else:
                    alvara.fase = None
                
                # Handle fase_honorarios_contratuais field properly (ForeignKey)
                fase_honorarios_id = request.POST.get('fase_honorarios_contratuais')
                if fase_honorarios_id:
                    try:
                        fase_honorarios = FaseHonorariosContratuais.objects.get(id=fase_honorarios_id)
                        alvara.fase_honorarios_contratuais = fase_honorarios
                    except FaseHonorariosContratuais.DoesNotExist:
                        messages.error(request, 'Fase Honorários Contratuais selecionada não existe.')
                        return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                else:
                    alvara.fase_honorarios_contratuais = None
                
                # Handle fase_honorarios_sucumbenciais field properly (ForeignKey)
                fase_honorarios_sucumbenciais_id = request.POST.get('fase_honorarios_sucumbenciais')
                if fase_honorarios_sucumbenciais_id:
                    try:
                        fase_honorarios_sucumbenciais = FaseHonorariosSucumbenciais.objects.get(id=fase_honorarios_sucumbenciais_id)
                        alvara.fase_honorarios_sucumbenciais = fase_honorarios_sucumbenciais
                    except FaseHonorariosSucumbenciais.DoesNotExist:
                        messages.error(request, 'Fase Honorários Sucumbenciais selecionada não existe.')
                        return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                else:
                    alvara.fase_honorarios_sucumbenciais = None
                
                alvara.save()
                
                # Check if any fase field was changed and add warning message
                fase_changed = (
                    original_fase != alvara.fase or
                    original_fase_honorarios_contratuais != alvara.fase_honorarios_contratuais or
                    original_fase_honorarios_sucumbenciais != alvara.fase_honorarios_sucumbenciais
                )
                
                messages.success(request, f'Alvará do cliente {alvara.cliente.nome} atualizado com sucesso!')
                
                if fase_changed:
                    # Instead of showing a warning message, we'll redirect with a flag to show modal
                    return redirect(f"{reverse('precatorio_detalhe', kwargs={'precatorio_cnj': precatorio.cnj})}?fase_changed=true")
                else:
                    # Regular redirect when no fase change occurred
                    return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)

            except Alvara.DoesNotExist:
                messages.error(request, 'Alvará não encontrado.')
            except (ValueError, TypeError) as e:
                messages.error(request, f'Erro nos dados fornecidos: {str(e)}')
            except Exception as e:
                messages.error(request, f'Erro inesperado: {str(e)}')
        
        elif 'delete_alvara' in request.POST:
            # Handle alvara deletion
            alvara_id = request.POST.get('alvara_id')
            try:
                alvara = get_object_or_404(Alvara, id=alvara_id, precatorio=precatorio)
                cliente_nome = alvara.cliente.nome
                alvara.delete()
                messages.success(request, f'Alvará do cliente {cliente_nome} excluído com sucesso!')
                return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            except Alvara.DoesNotExist:
                messages.error(request, 'Alvará não encontrado.')
        
        elif 'update_requerimento' in request.POST:
            # Handle requerimento update
            requerimento_id = request.POST.get('requerimento_id')
            try:
                requerimento = get_object_or_404(Requerimento, id=requerimento_id, precatorio=precatorio)
                # Update the requerimento fields
                requerimento.valor = float(request.POST.get('valor', requerimento.valor))
                requerimento.desagio = float(request.POST.get('desagio', requerimento.desagio))
                
                # Handle pedido field properly (ForeignKey)
                pedido_id = request.POST.get('pedido')
                if pedido_id:
                    try:
                        pedido = PedidoRequerimento.objects.get(id=pedido_id)
                        requerimento.pedido = pedido
                    except PedidoRequerimento.DoesNotExist:
                        messages.error(request, 'Tipo de pedido selecionado não existe.')
                        return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                else:
                    requerimento.pedido = None
                
                # Handle fase field properly (ForeignKey)
                fase_id = request.POST.get('fase')
                if fase_id:
                    try:
                        fase = Fase.objects.get(id=fase_id)
                        requerimento.fase = fase
                    except Fase.DoesNotExist:
                        messages.error(request, 'Fase selecionada não existe.')
                        return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
                else:
                    requerimento.fase = None
                
                requerimento.save()
                messages.success(request, f'Requerimento do cliente {requerimento.cliente.nome} atualizado com sucesso!')
                return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            except Requerimento.DoesNotExist:
                messages.error(request, 'Requerimento não encontrado.')
            except (ValueError, TypeError) as e:
                messages.error(request, f'Erro nos dados fornecidos: {str(e)}')
            except Exception as e:
                messages.error(request, f'Erro inesperado: {str(e)}')
        
        elif 'delete_requerimento' in request.POST:
            # Handle requerimento deletion
            requerimento_id = request.POST.get('requerimento_id')
            try:
                requerimento = get_object_or_404(Requerimento, id=requerimento_id, precatorio=precatorio)
                cliente_nome = requerimento.cliente.nome
                requerimento.delete()
                messages.success(request, f'Requerimento do cliente {cliente_nome} excluído com sucesso!')
                return redirect('precatorio_detalhe', precatorio_cnj=precatorio.cnj)
            except Requerimento.DoesNotExist:
                messages.error(request, 'Requerimento não encontrado.')
    else:
        # Initialize precatorio form for editing if requested
        if 'edit' in request.GET:
            precatorio_form = PrecatorioForm(instance=precatorio)
    
    context = {
        'precatorio': precatorio,
        'form': precatorio_form,
        'client_search_form': client_search_form,
        'cliente_form': cliente_form,
        'alvara_form': alvara_form,
        'requerimento_form': requerimento_form,
        'is_editing': request.method == 'POST' and 'edit_precatorio' in request.POST or 'edit' in request.GET,
        'edit_mode': 'edit' in request.GET,  # Safe boolean for template use
        'clientes': associated_clientes,
        'associated_clientes': associated_clientes,
        'alvaras': alvaras,
        'alvaras_total_principal': alvaras_total_principal,
        'alvaras_total_contratuais': alvaras_total_contratuais,
        'alvaras_total_sucumbenciais': alvaras_total_sucumbenciais,
        'requerimentos': requerimentos,
        'alvara_fases': Fase.get_fases_for_alvara(),
        'requerimento_fases': Fase.get_fases_for_requerimento(),
        'fases_honorarios_contratuais': FaseHonorariosContratuais.objects.filter(ativa=True),  # Uses model's default ordering: ['ordem', 'nome']
        'fases_honorarios_sucumbenciais': FaseHonorariosSucumbenciais.objects.filter(ativa=True),  # Uses model's default ordering: ['ordem', 'nome']
        'available_pedidos': PedidoRequerimento.get_ativos(),
        'contas_bancarias': ContaBancaria.objects.all().order_by('banco', 'agencia'),
    }
    
    return render(request, 'precapp/precatorio_detail.html', context)


@login_required
def delete_precatorio_view(request, precatorio_cnj):
    """View to delete a precatório"""
    precatorio = get_object_or_404(Precatorio, cnj=precatorio_cnj)
    
    if request.method == 'POST':
        precatorio_cnj_display = precatorio.cnj
        
        # Check if precatorio has associated clientes
        if precatorio.clientes.exists():
            messages.error(request, f'Não é possível excluir o precatório {precatorio_cnj_display} pois ele possui clientes associados. Remova as associações primeiro.')
            return redirect('precatorio_detalhe', precatorio_cnj=precatorio_cnj)
        
        # Check if precatorio has associated alvaras
        if hasattr(precatorio, 'alvara_set') and precatorio.alvara_set.exists():
            messages.error(request, f'Não é possível excluir o precatório {precatorio_cnj_display} pois ele possui alvarás associados. Remova os alvarás primeiro.')
            return redirect('precatorio_detalhe', precatorio_cnj=precatorio_cnj)
        
        # Check if precatorio has associated requerimentos
        if hasattr(precatorio, 'requerimento_set') and precatorio.requerimento_set.exists():
            messages.error(request, f'Não é possível excluir o precatório {precatorio_cnj_display} pois ele possui requerimentos associados. Remova os requerimentos primeiro.')
            return redirect('precatorio_detalhe', precatorio_cnj=precatorio_cnj)
        
        precatorio.delete()
        messages.success(request, f'Precatório {precatorio_cnj_display} foi excluído com sucesso!')
        return redirect('precatorios')
    
    # If not POST, redirect to precatorio detail
    return redirect('precatorio_detalhe', precatorio_cnj=precatorio_cnj)


# ===============================
# CLIENTE VIEWS
# ===============================


@login_required
def clientes_view(request):
    """View to display all clients with filtering support"""
    clientes = Cliente.objects.all().prefetch_related(
        'precatorios', 
        'precatorios__requerimento_set', 
        'precatorios__requerimento_set__fase'
    )
    
    # Get pedido names dynamically  
    prioridade_pedido_names = get_prioridade_pedido_names()
    
    # Apply filters based on GET parameters
    nome_filter = request.GET.get('nome', '').strip()
    cpf_filter = request.GET.get('cpf', '').strip()
    idade_filter = request.GET.get('idade', '').strip()
    prioridade_filter = request.GET.get('prioridade', '')
    requerimento_prioridade_filter = request.GET.get('requerimento_prioridade', '')
    precatorio_filter = request.GET.get('precatorio', '').strip()
    falecido_filter = request.GET.get('falecido', '')
    
    if nome_filter:
        clientes = clientes.filter(nome__icontains=nome_filter)
    
    if cpf_filter:
        clientes = clientes.filter(cpf__icontains=cpf_filter)
    
    # Filter by age
    if idade_filter:
        try:
            idade = int(idade_filter)
            from datetime import date
            from dateutil.relativedelta import relativedelta
            
            # Calculate the birth year range for clients of the specified age
            today = date.today()
            # For someone to be X years old today, they must have been born between:
            # - (today - X+1 years + 1 day) and (today - X years)
            min_birth_date = today - relativedelta(years=idade+1, days=-1)
            max_birth_date = today - relativedelta(years=idade)
            
            clientes = clientes.filter(
                nascimento__gte=min_birth_date,
                nascimento__lte=max_birth_date
            )
        except ValueError:
            # If idade_filter is not a valid integer, ignore the filter
            pass

    if prioridade_filter in ['true', 'false']:
        prioridade_bool = prioridade_filter == 'true'
        clientes = clientes.filter(prioridade=prioridade_bool)
    
    if falecido_filter in ['true', 'false']:
        falecido_bool = falecido_filter == 'true'
        clientes = clientes.filter(falecido=falecido_bool)
    
    # Filter by requerimento prioridade (based on Deferido/Não Deferido status)
    if requerimento_prioridade_filter:
        if requerimento_prioridade_filter == 'deferido':
            # Find clientes that have priority requerimentos with 'Deferido' phase
            clientes_deferidos = Requerimento.objects.filter(
                pedido__nome__in=prioridade_pedido_names,
                fase__nome='Deferido'
            ).values_list('cliente__cpf', flat=True).distinct()
            clientes = clientes.filter(cpf__in=clientes_deferidos)
        elif requerimento_prioridade_filter == 'nao_deferido':
            # Find clientes that have priority requerimentos that are NOT 'Deferido'
            clientes_nao_deferidos = Requerimento.objects.filter(
                pedido__nome__in=prioridade_pedido_names
            ).exclude(
                fase__nome='Deferido'
            ).values_list('cliente__cpf', flat=True).distinct()
            clientes = clientes.filter(cpf__in=clientes_nao_deferidos)
        elif requerimento_prioridade_filter == 'sem_requerimento':
            # Find clientes that have NO priority requerimentos at all
            clientes_com_requerimentos = Requerimento.objects.filter(
                pedido__nome__in=prioridade_pedido_names
            ).values_list('cliente__cpf', flat=True).distinct()
            clientes = clientes.exclude(cpf__in=clientes_com_requerimentos)
    
    if precatorio_filter:
        clientes = clientes.filter(precatorios__cnj__icontains=precatorio_filter).distinct()
    
    # Calculate summary statistics
    total_clientes = clientes.count()
    clientes_com_prioridade = clientes.filter(prioridade=True).count()
    clientes_sem_prioridade = clientes.filter(prioridade=False).count()
    
    context = {
        'clientes': clientes,
        'total_clientes': total_clientes,
        'clientes_com_prioridade': clientes_com_prioridade,
        'clientes_sem_prioridade': clientes_sem_prioridade,
        # Include current filter values to maintain state in form
        'current_nome': nome_filter,
        'current_cpf': cpf_filter,
        'current_idade': idade_filter,
        'current_prioridade': prioridade_filter,
        'current_requerimento_prioridade': requerimento_prioridade_filter,
        'current_precatorio': precatorio_filter,
        'current_falecido': falecido_filter,
    }
    
    return render(request, 'precapp/cliente_list.html', context)


@login_required
def cliente_detail_view(request, cpf):
    """View to display and edit a specific client and manage precatorio relationships"""
    cliente = get_object_or_404(Cliente, cpf=cpf)
    
    # Get all precatorios associated with this client
    associated_precatorios = cliente.precatorios.all().order_by('cnj')
    
    # Get all diligencias for this client
    diligencias = cliente.diligencias.all().select_related('tipo', 'responsavel').order_by('-data_criacao')
    diligencias_pendentes = diligencias.filter(concluida=False)
    diligencias_concluidas = diligencias.filter(concluida=True)
    
    # Initialize forms
    client_form = None
    search_form = PrecatorioSearchForm()
    
    if request.method == 'POST':
        if 'edit_client' in request.POST:
            # Handle client editing
            client_form = ClienteForm(request.POST, instance=cliente)
            if client_form.is_valid():
                client_form.save()
                messages.success(request, f'Cliente {cliente.nome} atualizado com sucesso!')
                return redirect('cliente_detail', cpf=cliente.cpf)
            else:
                messages.error(request, 'Por favor, corrija os erros abaixo.')
        
        elif 'link_precatorio' in request.POST:
            # Handle precatorio linking
            search_form = PrecatorioSearchForm(request.POST)
            if search_form.is_valid():
                cnj = search_form.cleaned_data['cnj']
                try:
                    precatorio = Precatorio.objects.get(cnj=cnj)
                    if precatorio in associated_precatorios:
                        messages.warning(request, f'O cliente {cliente.nome} já está vinculado ao precatório {cnj}.')
                    else:
                        # Add the client to the precatorio
                        precatorio.clientes.add(cliente)
                        messages.success(request, f'Cliente {cliente.nome} vinculado com sucesso ao precatório {cnj}!')
                        return redirect('cliente_detail', cpf=cliente.cpf)
                except Precatorio.DoesNotExist:
                    messages.error(request, f'Precatório com CNJ {cnj} não encontrado. Verifique se o número está correto.')
            else:
                messages.error(request, 'Por favor, corrija os erros no CNJ.')
        
        elif 'unlink_precatorio' in request.POST:
            # Handle precatorio unlinking
            precatorio_cnj = request.POST.get('precatorio_cnj')
            try:
                precatorio = Precatorio.objects.get(cnj=precatorio_cnj)
                if precatorio in associated_precatorios:
                    precatorio.clientes.remove(cliente)
                    messages.success(request, f'Cliente {cliente.nome} desvinculado do precatório {precatorio.cnj}.')
                    return redirect('cliente_detail', cpf=cliente.cpf)
                else:
                    messages.error(request, 'Este cliente não está vinculado a este precatório.')
            except Precatorio.DoesNotExist:
                messages.error(request, 'Precatório não encontrado.')
    else:
        # Initialize client form for editing if requested
        if 'edit' in request.GET:
            client_form = ClienteForm(instance=cliente)
    
    context = {
        'cliente': cliente,
        'client_form': client_form,
        'search_form': search_form,
        'associated_precatorios': associated_precatorios,
        'is_editing': request.method == 'POST' and 'edit_client' in request.POST or 'edit' in request.GET,
        'diligencias': diligencias,
        'diligencias_pendentes': diligencias_pendentes,
        'diligencias_concluidas': diligencias_concluidas,
        'total_diligencias': diligencias.count(),
        'total_pendentes': diligencias_pendentes.count(),
        'total_concluidas': diligencias_concluidas.count(),
    }
    
    return render(request, 'precapp/cliente_detail.html', context)


@login_required
def novo_cliente_view(request):
    """View to create a new client"""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            messages.success(request, f'Cliente {cliente.nome} foi criado com sucesso!')
            return redirect('cliente_detail', cpf=cliente.cpf)
    else:
        form = ClienteForm()
    
    context = {
        'form': form,
    }
    return render(request, 'precapp/novo_cliente.html', context)


@login_required
def delete_cliente_view(request, cpf):
    """View to delete a client"""
    cliente = get_object_or_404(Cliente, cpf=cpf)
    
    if request.method == 'POST':
        cliente_nome = cliente.nome
        
        # Check if cliente has associated precatorios
        if cliente.precatorios.exists():
            messages.error(request, f'Não é possível excluir o cliente {cliente_nome} pois ele está associado a precatórios. Remova as associações primeiro.')
            return redirect('cliente_detail', cpf=cpf)
        
        # Check if cliente has associated alvaras
        if hasattr(cliente, 'alvara_set') and cliente.alvara_set.exists():
            messages.error(request, f'Não é possível excluir o cliente {cliente_nome} pois ele possui alvarás associados. Remova os alvarás primeiro.')
            return redirect('cliente_detail', cpf=cpf)
        
        # Check if cliente has associated requerimentos
        if hasattr(cliente, 'requerimento_set') and cliente.requerimento_set.exists():
            messages.error(request, f'Não é possível excluir o cliente {cliente_nome} pois ele possui requerimentos associados. Remova os requerimentos primeiro.')
            return redirect('cliente_detail', cpf=cpf)
        
        cliente.delete()
        messages.success(request, f'Cliente {cliente_nome} foi excluído com sucesso!')
        return redirect('clientes')
    
    # If not POST, redirect to client detail
    return redirect('cliente_detail', cpf=cpf)


@login_required
@require_http_methods(["POST"])
def update_priority_by_age(request):
    """Update priority status for clients over 60 years old"""
    from datetime import date, timedelta
    
    try:
        # Calculate date 60 years ago
        sixty_years_ago = date.today() - timedelta(days=60*365.25)
        
        # Find clients born before this date (older than 60) without priority
        clients_over_60 = Cliente.objects.filter(
            nascimento__lt=sixty_years_ago,
            prioridade=False
        )
        
        count_before = clients_over_60.count()
        
        if count_before > 0:
            # Update their priority status
            updated_count = clients_over_60.update(prioridade=True)
            
            messages.success(
                request, 
                f'✅ {updated_count} cliente(s) com mais de 60 anos foram atualizados para status prioritário.'
            )
            
            # Log some examples for transparency
            if updated_count > 0:
                examples = Cliente.objects.filter(
                    nascimento__lt=sixty_years_ago,
                    prioridade=True
                )[:3]
                
                example_names = [client.nome for client in examples]
                if len(example_names) > 0:
                    examples_text = ", ".join(example_names)
                    if updated_count > 3:
                        examples_text += f" e mais {updated_count - 3} cliente(s)"
                    
                    messages.info(
                        request,
                        f'Exemplos atualizados: {examples_text}'
                    )
        else:
            messages.info(
                request,
                '📋 Nenhum cliente encontrado que precise ser atualizado. Todos os clientes com mais de 60 anos já possuem status prioritário.'
            )
            
    except Exception as e:
        messages.error(
            request,
            f'❌ Erro ao atualizar prioridades: {str(e)}'
        )
    
    return redirect('clientes')


# ===============================
# ALVARA VIEWS
# ===============================


@login_required
def alvaras_view(request):
    """View to display all alvarás with filtering support"""
    alvaras = Alvara.objects.all().select_related(
        'precatorio', 'cliente', 'fase', 'fase_honorarios_contratuais', 'fase_honorarios_sucumbenciais'
    ).order_by('-id')
    
    # Get available fases for alvara
    available_fases = Fase.get_fases_for_alvara()
    # Get available fases for honorários contratuais
    available_fases_honorarios = FaseHonorariosContratuais.objects.all()  # Uses model's default ordering: ['ordem', 'nome']
    # Get available fases for honorários sucumbenciais
    available_fases_honorarios_sucumbenciais = FaseHonorariosSucumbenciais.objects.all()  # Uses model's default ordering: ['ordem', 'nome']
    
    # Apply filters based on GET parameters
    nome_filter = request.GET.get('nome', '').strip()
    precatorio_filter = request.GET.get('precatorio', '').strip()
    tipo_filter = request.GET.get('tipo', '').strip()
    fase_filter = request.GET.get('fase', '').strip()
    fase_honorarios_filter = request.GET.get('fase_honorarios', '').strip()
    fase_honorarios_sucumbenciais_filter = request.GET.get('fase_honorarios_sucumbenciais', '').strip()
    
    if nome_filter:
        alvaras = alvaras.filter(cliente__nome__icontains=nome_filter)
    
    if precatorio_filter:
        alvaras = alvaras.filter(precatorio__cnj__icontains=precatorio_filter)
    
    if tipo_filter:
        alvaras = alvaras.filter(tipo=tipo_filter)  # Exact match for dropdown
    
    if fase_filter:
        alvaras = alvaras.filter(fase__nome=fase_filter)  # Exact match for dropdown
        
    if fase_honorarios_filter:
        alvaras = alvaras.filter(fase_honorarios_contratuais__nome=fase_honorarios_filter)  # Exact match for dropdown
    
    if fase_honorarios_sucumbenciais_filter:
        alvaras = alvaras.filter(fase_honorarios_sucumbenciais__nome=fase_honorarios_sucumbenciais_filter)  # Exact match for dropdown
    
    # Calculate summary statistics
    total_alvaras = alvaras.count()
    aguardando_deposito = alvaras.filter(tipo='aguardando depósito').count()
    deposito_judicial = alvaras.filter(tipo='depósito judicial').count()
    recebido_cliente = alvaras.filter(tipo='recebido pelo cliente').count()
    honorarios_recebidos = alvaras.filter(tipo='honorários recebidos').count()
    
    # Calculate total values
    total_valor_principal = sum([alvara.valor_principal for alvara in alvaras if alvara.valor_principal])
    total_honorarios_contratuais = sum([alvara.honorarios_contratuais for alvara in alvaras if alvara.honorarios_contratuais])
    total_honorarios_sucumbenciais = sum([alvara.honorarios_sucumbenciais for alvara in alvaras if alvara.honorarios_sucumbenciais])
    
    context = {
        'alvaras': alvaras,
        'total_alvaras': total_alvaras,
        'aguardando_deposito': aguardando_deposito,
        'deposito_judicial': deposito_judicial,
        'recebido_cliente': recebido_cliente,
        'honorarios_recebidos': honorarios_recebidos,
        'total_valor_principal': total_valor_principal,
        'total_honorarios_contratuais': total_honorarios_contratuais,
        'total_honorarios_sucumbenciais': total_honorarios_sucumbenciais,
        # Include current filter values to maintain state in form
        'current_nome': nome_filter,
        'current_precatorio': precatorio_filter,
        'current_tipo': tipo_filter,
        'current_fase': fase_filter,
        'current_fase_honorarios': fase_honorarios_filter,
        'current_fase_honorarios_sucumbenciais': fase_honorarios_sucumbenciais_filter,
        # Include available options for dropdowns
        'available_fases': available_fases,
        'available_fases_honorarios': available_fases_honorarios,
        'available_fases_honorarios_sucumbenciais': available_fases_honorarios_sucumbenciais,
    }
    
    return render(request, 'precapp/alvara_list.html', context)


@login_required
def delete_alvara_view(request, alvara_id):
    """View to delete a specific alvará"""
    alvara = get_object_or_404(Alvara, id=alvara_id)
    
    if request.method == 'POST':
        precatorio_cnj = alvara.precatorio.cnj
        cliente_nome = alvara.cliente.nome
        alvara.delete()
        messages.success(request, f'Alvará de {cliente_nome} foi excluído com sucesso!')
        return redirect('alvaras')
    
    # If GET request, redirect back to alvaras list instead of detail (which doesn't exist)
    return redirect('alvaras')


# ===============================
# REQUERIMENTO VIEWS
# ===============================


@login_required
def requerimento_list_view(request):
    """View to list all requerimentos with filtering"""
    requerimentos = Requerimento.objects.all().select_related(
        'cliente', 'precatorio', 'fase', 'pedido'
    ).order_by('-id')
    
    # Get filter parameters
    cliente_filter = request.GET.get('cliente', '').strip()
    precatorio_filter = request.GET.get('precatorio', '').strip()
    cnj_requerimento_filter = request.GET.get('cnj_requerimento', '').strip()
    pedido_filter = request.GET.get('pedido', '').strip()
    fase_filter = request.GET.get('fase', '').strip()
    
    # Apply filters
    if cliente_filter:
        requerimentos = requerimentos.filter(cliente__nome__icontains=cliente_filter)

    if precatorio_filter:
        requerimentos = requerimentos.filter(precatorio__cnj__icontains=precatorio_filter)

    if cnj_requerimento_filter:
        requerimentos = requerimentos.filter(cnj__icontains=cnj_requerimento_filter)

    if pedido_filter:
        requerimentos = requerimentos.filter(pedido__id=pedido_filter)

    if fase_filter:
        requerimentos = requerimentos.filter(fase__nome=fase_filter)    # Get available phases for requerimentos
    from .models import Fase
    available_fases = Fase.get_fases_for_requerimento()
    
    # Get available pedido requerimento types
    available_pedidos = PedidoRequerimento.get_ativos()
    
    # Calculate financial statistics based on filtered results
    valor_total = sum(r.valor for r in requerimentos if r.valor)
    desagio_medio = sum(r.desagio for r in requerimentos if r.desagio) / requerimentos.count() if requerimentos.count() > 0 else 0
    
    # Handle requerimento deletion
    if request.method == 'POST' and 'delete_requerimento' in request.POST:
        requerimento_id = request.POST.get('requerimento_id')
        try:
            requerimento = Requerimento.objects.get(id=requerimento_id)
            cliente_nome = requerimento.cliente.nome
            requerimento.delete()
            messages.success(request, f'Requerimento de {cliente_nome} foi excluído com sucesso!')
            return redirect('requerimentos')
        except Requerimento.DoesNotExist:
            messages.error(request, 'Requerimento não encontrado.')
    
    context = {
        'requerimentos': requerimentos,
        'available_fases': available_fases,
        'available_pedidos': available_pedidos,
        'valor_total': valor_total,
        'desagio_medio': desagio_medio,
        # Current filter values for form persistence
        'current_cliente': cliente_filter,
        'current_precatorio': precatorio_filter,
        'current_cnj_requerimento': cnj_requerimento_filter,
        'current_pedido': pedido_filter,
        'current_fase': fase_filter,
    }
    return render(request, 'precapp/requerimento_list.html', context)


# ===============================
# FASE MANAGEMENT VIEWS
# ===============================

@login_required
def fases_view(request):
    """View to list all phases"""
    fases = Fase.objects.all()  # Uses model's default ordering: ['ordem', 'tipo', 'nome']
    
    # Statistics
    total_fases = fases.count()
    fases_ativas = fases.filter(ativa=True).count()
    fases_inativas = fases.filter(ativa=False).count()
    
    context = {
        'fases': fases,
        'total_fases': total_fases,
        'fases_ativas': fases_ativas,
        'fases_inativas': fases_inativas,
    }
    
    return render(request, 'precapp/fases_list.html', context)


@login_required
def nova_fase_view(request):
    """View to create a new phase"""
    if request.method == 'POST':
        form = FaseForm(request.POST)
        if form.is_valid():
            fase = form.save()
            messages.success(request, f'Fase "{fase.nome}" criada com sucesso!')
            return redirect('fases')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = FaseForm()
    
    context = {
        'form': form,
        'title': 'Nova Fase',
        'submit_text': 'Criar Fase'
    }
    return render(request, 'precapp/fase_form.html', context)


@login_required
def editar_fase_view(request, fase_id):
    """View to edit an existing phase"""
    fase = get_object_or_404(Fase, id=fase_id)
    
    if request.method == 'POST':
        form = FaseForm(request.POST, instance=fase)
        if form.is_valid():
            fase = form.save()
            messages.success(request, f'Fase "{fase.nome}" atualizada com sucesso!')
            return redirect('fases')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = FaseForm(instance=fase)
    
    context = {
        'form': form,
        'fase': fase,
        'title': f'Editar Fase: {fase.nome}',
        'submit_text': 'Salvar Alterações'
    }
    return render(request, 'precapp/fase_form.html', context)


@login_required
def deletar_fase_view(request, fase_id):
    """View to delete a phase"""
    fase = get_object_or_404(Fase, id=fase_id)
    
    if request.method == 'POST':
        fase_nome = fase.nome
        
        # Check if fase is being used by any alvara
        alvaras_using_fase = Alvara.objects.filter(fase=fase)
        if alvaras_using_fase.exists():
            messages.error(
                request, 
                f'Não é possível excluir a fase "{fase_nome}" pois ela está sendo usada por {alvaras_using_fase.count()} alvará(s). '
                'Altere a fase desses alvarás primeiro.'
            )
            return redirect('fases')
        
        # Check if fase is being used by any requerimento
        requerimentos_using_fase = Requerimento.objects.filter(fase=fase)
        if requerimentos_using_fase.exists():
            messages.error(
                request, 
                f'Não é possível excluir a fase "{fase_nome}" pois ela está sendo usada por {requerimentos_using_fase.count()} requerimento(s). '
                'Altere a fase desses requerimentos primeiro.'
            )
            return redirect('fases')
        
        fase.delete()
        messages.success(request, f'Fase "{fase_nome}" foi excluída com sucesso!')
        return redirect('fases')
    
    # If not POST, redirect to fases list
    return redirect('fases')


@login_required
def ativar_fase_view(request, fase_id):
    """View to activate/deactivate a phase"""
    fase = get_object_or_404(Fase, id=fase_id)
    
    if request.method == 'POST':
        fase.ativa = not fase.ativa
        fase.save()
        
        status_text = "ativada" if fase.ativa else "desativada"
        messages.success(request, f'Fase "{fase.nome}" foi {status_text} com sucesso!')
    
    return redirect('fases')



# ===============================
# FASE HONORÁRIOS CONTRATUAIS VIEWS
# ===============================

@login_required
def fases_honorarios_view(request):
    """View to list all honorários contratuais phases"""
    fases = FaseHonorariosContratuais.objects.all()  # Uses model's default ordering: ['ordem', 'nome']
    
    # Statistics
    total_fases = fases.count()
    fases_ativas = fases.filter(ativa=True).count()
    fases_inativas = fases.filter(ativa=False).count()
    
    context = {
        'fases': fases,
        'total_fases': total_fases,
        'fases_ativas': fases_ativas,
        'fases_inativas': fases_inativas,
    }
    
    return render(request, 'precapp/fases_honorarios_list.html', context)


@login_required
def nova_fase_honorarios_view(request):
    """View to create a new honorários contratuais phase"""
    if request.method == 'POST':
        form = FaseHonorariosContratuaisForm(request.POST)
        if form.is_valid():
            fase = form.save()
            messages.success(request, f'Fase de Honorários Contratuais "{fase.nome}" criada com sucesso!')
            return redirect('fases_honorarios')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = FaseHonorariosContratuaisForm()
    
    context = {
        'form': form,
        'title': 'Nova Fase Honorários Contratuais',
        'submit_text': 'Criar Fase'
    }
    return render(request, 'precapp/fase_honorarios_form.html', context)


@login_required
def editar_fase_honorarios_view(request, fase_id):
    """View to edit an existing honorários contratuais phase"""
    fase = get_object_or_404(FaseHonorariosContratuais, id=fase_id)
    
    if request.method == 'POST':
        form = FaseHonorariosContratuaisForm(request.POST, instance=fase)
        if form.is_valid():
            fase = form.save()
            messages.success(request, f'Fase de Honorários Contratuais "{fase.nome}" atualizada com sucesso!')
            return redirect('fases_honorarios')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = FaseHonorariosContratuaisForm(instance=fase)
    
    context = {
        'form': form,
        'fase': fase,
        'title': f'Editar Fase Honorários Contratuais: {fase.nome}',
        'submit_text': 'Salvar Alterações'
    }
    return render(request, 'precapp/fase_honorarios_form.html', context)


@login_required
def deletar_fase_honorarios_view(request, fase_id):
    """View to delete a honorários contratuais phase"""
    fase = get_object_or_404(FaseHonorariosContratuais, id=fase_id)
    
    if request.method == 'POST':
        fase_nome = fase.nome
        
        # Check if fase is being used by any alvara
        alvaras_using_fase = Alvara.objects.filter(fase_honorarios_contratuais=fase)
        if alvaras_using_fase.exists():
            messages.error(
                request, 
                f'Não é possível excluir a fase de honorários contratuais "{fase_nome}" pois ela está sendo usada por {alvaras_using_fase.count()} alvará(s). '
                'Altere a fase desses alvarás primeiro.'
            )
            return redirect('fases_honorarios')
        
        fase.delete()
        messages.success(request, f'Fase de Honorários Contratuais "{fase_nome}" foi excluída com sucesso!')
        return redirect('fases_honorarios')
    
    # If not POST, redirect to fases list
    return redirect('fases_honorarios')


@login_required
def ativar_fase_honorarios_view(request, fase_id):
    """View to activate/deactivate a honorários contratuais phase"""
    fase = get_object_or_404(FaseHonorariosContratuais, id=fase_id)
    
    if request.method == 'POST':
        fase.ativa = not fase.ativa
        fase.save()
        
        status_text = "ativada" if fase.ativa else "desativada"
        messages.success(request, f'Fase de Honorários Contratuais "{fase.nome}" foi {status_text} com sucesso!')
    
    return redirect('fases_honorarios')


# ===============================
# FASE HONORÁRIOS SUCUMBENCIAIS VIEWS
# ===============================

@login_required
def fases_honorarios_sucumbenciais_view(request):
    """View to list all honorários sucumbenciais phases"""
    fases = FaseHonorariosSucumbenciais.objects.all()  # Uses model's default ordering: ['ordem', 'nome']
    
    # Statistics
    total_fases = fases.count()
    fases_ativas = fases.filter(ativa=True).count()
    fases_inativas = fases.filter(ativa=False).count()
    
    context = {
        'fases': fases,
        'total_fases': total_fases,
        'fases_ativas': fases_ativas,
        'fases_inativas': fases_inativas,
    }
    
    return render(request, 'precapp/fases_honorarios_sucumbenciais_list.html', context)


@login_required
def nova_fase_honorarios_sucumbenciais_view(request):
    """View to create a new honorários sucumbenciais phase"""
    if request.method == 'POST':
        form = FaseHonorariosSucumbenciaisForm(request.POST)
        if form.is_valid():
            fase = form.save()
            messages.success(request, f'Fase de Honorários Sucumbenciais "{fase.nome}" criada com sucesso!')
            return redirect('fases_honorarios_sucumbenciais')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = FaseHonorariosSucumbenciaisForm()
    
    context = {
        'form': form,
        'title': 'Nova Fase Honorários Sucumbenciais',
        'submit_text': 'Criar Fase'
    }
    return render(request, 'precapp/fase_honorarios_sucumbenciais_form.html', context)


@login_required
def editar_fase_honorarios_sucumbenciais_view(request, fase_id):
    """View to edit an existing honorários sucumbenciais phase"""
    fase = get_object_or_404(FaseHonorariosSucumbenciais, id=fase_id)
    
    if request.method == 'POST':
        form = FaseHonorariosSucumbenciaisForm(request.POST, instance=fase)
        if form.is_valid():
            fase = form.save()
            messages.success(request, f'Fase de Honorários Sucumbenciais "{fase.nome}" atualizada com sucesso!')
            return redirect('fases_honorarios_sucumbenciais')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = FaseHonorariosSucumbenciaisForm(instance=fase)
    
    context = {
        'form': form,
        'fase': fase,
        'title': f'Editar Fase Honorários Sucumbenciais: {fase.nome}',
        'submit_text': 'Salvar Alterações'
    }
    return render(request, 'precapp/fase_honorarios_sucumbenciais_form.html', context)


@login_required
def deletar_fase_honorarios_sucumbenciais_view(request, fase_id):
    """View to delete a honorários sucumbenciais phase"""
    fase = get_object_or_404(FaseHonorariosSucumbenciais, id=fase_id)
    
    if request.method == 'POST':
        fase_nome = fase.nome
        
        # Check if fase is being used by any alvara
        alvaras_using_fase = Alvara.objects.filter(fase_honorarios_sucumbenciais=fase)
        if alvaras_using_fase.exists():
            messages.error(
                request, 
                f'Não é possível excluir a fase de honorários sucumbenciais "{fase_nome}" pois ela está sendo usada por {alvaras_using_fase.count()} alvará(s). '
                'Altere a fase desses alvarás primeiro.'
            )
            return redirect('fases_honorarios_sucumbenciais')
        
        fase.delete()
        messages.success(request, f'Fase de Honorários Sucumbenciais "{fase_nome}" foi excluída com sucesso!')
        return redirect('fases_honorarios_sucumbenciais')
    
    # If not POST, redirect to fases list
    return redirect('fases_honorarios_sucumbenciais')


@login_required
def ativar_fase_honorarios_sucumbenciais_view(request, fase_id):
    """View to activate/deactivate a honorários sucumbenciais phase"""
    fase = get_object_or_404(FaseHonorariosSucumbenciais, id=fase_id)
    
    if request.method == 'POST':
        fase.ativa = not fase.ativa
        fase.save()
        
        status_text = "ativada" if fase.ativa else "desativada"
        messages.success(request, f'Fase de Honorários Sucumbenciais "{fase.nome}" foi {status_text} com sucesso!')
    
    return redirect('fases_honorarios_sucumbenciais')


# ===============================
# TIPO DILIGENCIA VIEWS
# ===============================

@login_required
def tipos_diligencia_view(request):
    """View to list all diligence types"""
    tipos = TipoDiligencia.objects.all()  # Uses model's default ordering: ['ordem', 'nome']
    
    # Add diligencias count for each tipo
    for tipo in tipos:
        tipo.diligencias_count = tipo.diligencias_set.count()
    
    # Statistics
    total_tipos = tipos.count()
    tipos_ativos = tipos.filter(ativo=True).count()
    tipos_inativos = tipos.filter(ativo=False).count()
    
    # Get count of total diligencias using the reverse relationship
    total_diligencias = sum(tipo.diligencias_set.count() for tipo in tipos)
    
    context = {
        'tipos_diligencia': tipos,  # Changed from 'tipos' to match template
        'total_tipos': total_tipos,
        'tipos_ativos': tipos_ativos,
        'tipos_inativos': tipos_inativos,
        'total_diligencias': total_diligencias,
    }
    
    return render(request, 'precapp/tipos_diligencia_list.html', context)


@login_required
def novo_tipo_diligencia_view(request):
    """View to create a new diligence type"""
    if request.method == 'POST':
        form = TipoDiligenciaForm(request.POST)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo de Diligência "{tipo.nome}" criado com sucesso!')
            return redirect('tipos_diligencia')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = TipoDiligenciaForm()
    
    context = {
        'form': form,
        'title': 'Novo Tipo de Diligência',
        'action': 'Criar'
    }
    
    return render(request, 'precapp/tipo_diligencia_form.html', context)


@login_required
def editar_tipo_diligencia_view(request, tipo_id):
    """View to edit an existing diligence type"""
    tipo = get_object_or_404(TipoDiligencia, id=tipo_id)
    
    if request.method == 'POST':
        # Check if this is an update from the dropdown form in list view
        if 'update_tipo' in request.POST:
            # Handle direct field updates from dropdown form
            nome = request.POST.get('nome')
            cor = request.POST.get('cor')
            descricao = request.POST.get('descricao', '')
            ativo = 'ativo' in request.POST
            
            if nome and cor:
                tipo.nome = nome
                tipo.cor = cor
                tipo.descricao = descricao
                tipo.ativo = ativo
                tipo.save()
                messages.success(request, f'Tipo de Diligência "{tipo.nome}" atualizado com sucesso!')
                return redirect('tipos_diligencia')
            else:
                messages.error(request, 'Nome e cor são campos obrigatórios.')
                return redirect('tipos_diligencia')
        else:
            # Handle regular form update
            form = TipoDiligenciaForm(request.POST, instance=tipo)
            if form.is_valid():
                tipo = form.save()
                messages.success(request, f'Tipo de Diligência "{tipo.nome}" atualizado com sucesso!')
                return redirect('tipos_diligencia')
            else:
                messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = TipoDiligenciaForm(instance=tipo)
    
    context = {
        'form': form,
        'tipo': tipo,
        'title': f'Editar Tipo: {tipo.nome}',
        'action': 'Atualizar'
    }
    
    return render(request, 'precapp/tipo_diligencia_form.html', context)


@login_required
def deletar_tipo_diligencia_view(request, tipo_id):
    """View to delete a diligence type"""
    tipo = get_object_or_404(TipoDiligencia, id=tipo_id)
    
    if request.method == 'POST':
        try:
            nome_tipo = tipo.nome
            tipo.delete()
            messages.success(request, f'Tipo de Diligência "{nome_tipo}" excluído com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao excluir tipo: {str(e)}')
        
        return redirect('tipos_diligencia')
    
    # Count related diligencias
    diligencias_count = tipo.diligencias.count() if hasattr(tipo, 'diligencias') else 0
    
    context = {
        'tipo': tipo,
        'diligencias_count': diligencias_count,
        'can_delete': diligencias_count == 0
    }
    
    return render(request, 'precapp/confirmar_delete_tipo_diligencia.html', context)


@login_required  
def ativar_tipo_diligencia_view(request, tipo_id):
    """View to toggle active status of a diligence type"""
    tipo = get_object_or_404(TipoDiligencia, id=tipo_id)
    
    if request.method == 'POST':
        # Handle POST from dropdown form
        ativo_value = request.POST.get('ativo', 'false')
        tipo.ativo = ativo_value.lower() == 'true'
    else:
        # Handle GET for backward compatibility
        ativo_param = request.GET.get('ativo', '').lower()
        if ativo_param == 'true':
            tipo.ativo = True
        elif ativo_param == 'false':
            tipo.ativo = False
        else:
            # Toggle if no parameter
            tipo.ativo = not tipo.ativo
    
    tipo.save()
    
    status = "ativado" if tipo.ativo else "desativado"
    messages.success(request, f'Tipo de Diligência "{tipo.nome}" {status} com sucesso!')
    
    return redirect('tipos_diligencia')


# ===============================
# TIPO PRECATÓRIO VIEWS
# ===============================

@login_required
def tipos_precatorio_view(request):
    """List all tipos de precatório"""
    tipos = Tipo.objects.all().order_by('ordem', 'nome')
    
    context = {
        'tipos': tipos,
        'total_tipos': tipos.count(),
        'tipos_ativos': tipos.filter(ativa=True).count(),
        'tipos_inativos': tipos.filter(ativa=False).count(),
    }
    
    return render(request, 'precapp/tipos_precatorio_list.html', context)


@login_required
def novo_tipo_precatorio_view(request):
    """Create a new tipo de precatório"""
    if request.method == 'POST':
        form = TipoForm(request.POST)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo "{tipo.nome}" criado com sucesso!')
            return redirect('tipos_precatorio')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = TipoForm()
    
    return render(request, 'precapp/tipo_precatorio_form.html', {
        'form': form,
        'title': 'Novo Tipo de Precatório'
    })


@login_required
def editar_tipo_precatorio_view(request, tipo_id):
    """Edit an existing tipo de precatório"""
    tipo = get_object_or_404(Tipo, id=tipo_id)
    
    if request.method == 'POST':
        form = TipoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo "{tipo.nome}" atualizado com sucesso!')
            return redirect('tipos_precatorio')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = TipoForm(instance=tipo)
    
    return render(request, 'precapp/tipo_precatorio_form.html', {
        'form': form,
        'tipo': tipo,
        'title': f'Editar Tipo: {tipo.nome}'
    })


@login_required
def deletar_tipo_precatorio_view(request, tipo_id):
    """Delete a tipo de precatório"""
    tipo = get_object_or_404(Tipo, id=tipo_id)
    
    # Check if tipo is being used
    precatorios_count = tipo.precatorio_set.count()
    
    if precatorios_count > 0:
        messages.error(
            request, 
            f'Não é possível excluir o tipo "{tipo.nome}" pois ele está sendo usado por {precatorios_count} precatório(s).'
        )
        return redirect('tipos_precatorio')
    
    if request.method == 'POST':
        tipo_nome = tipo.nome
        tipo.delete()
        messages.success(request, f'Tipo "{tipo_nome}" excluído com sucesso!')
        return redirect('tipos_precatorio')
    
    return render(request, 'precapp/confirmar_delete_tipo_precatorio.html', {
        'tipo': tipo,
        'precatorios_count': precatorios_count
    })


@login_required
def ativar_tipo_precatorio_view(request, tipo_id):
    """Toggle activation status of a tipo de precatório"""
    tipo = get_object_or_404(Tipo, id=tipo_id)
    
    # Handle POST data (from AJAX calls)
    if request.method == 'POST':
        ativo_value = request.POST.get('ativo', '').strip()
        tipo.ativa = ativo_value.lower() == 'true'
    else:
        # Handle GET for backward compatibility
        ativo_param = request.GET.get('ativo', '').lower()
        if ativo_param == 'true':
            tipo.ativa = True
        elif ativo_param == 'false':
            tipo.ativa = False
        else:
            # Toggle if no parameter
            tipo.ativa = not tipo.ativa
    
    tipo.save()
    
    status = "ativado" if tipo.ativa else "desativado"
    messages.success(request, f'Tipo de Precatório "{tipo.nome}" {status} com sucesso!')
    
    return redirect('tipos_precatorio')


# ===============================
# PEDIDO REQUERIMENTO VIEWS
# ===============================

@login_required
def tipos_pedido_requerimento_view(request):
    """List all tipos de pedido de requerimento"""
    tipos = PedidoRequerimento.objects.all().order_by('ordem', 'nome')
    
    context = {
        'tipos': tipos,
        'total_tipos': tipos.count(),
        'tipos_ativos': tipos.filter(ativo=True).count(),
        'tipos_inativos': tipos.filter(ativo=False).count(),
    }
    
    return render(request, 'precapp/tipos_pedido_requerimento_list.html', context)


@login_required
def novo_tipo_pedido_requerimento_view(request):
    """Create a new tipo de pedido de requerimento"""
    if request.method == 'POST':
        form = PedidoRequerimentoForm(request.POST)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo de Pedido "{tipo.nome}" criado com sucesso!')
            return redirect('tipos_pedido_requerimento')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = PedidoRequerimentoForm()
    
    return render(request, 'precapp/tipo_pedido_requerimento_form.html', {
        'form': form,
        'title': 'Novo Tipo de Pedido de Requerimento'
    })


@login_required
def editar_tipo_pedido_requerimento_view(request, tipo_id):
    """Edit an existing tipo de pedido de requerimento"""
    tipo = get_object_or_404(PedidoRequerimento, id=tipo_id)
    
    if request.method == 'POST':
        form = PedidoRequerimentoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo de Pedido "{tipo.nome}" atualizado com sucesso!')
            return redirect('tipos_pedido_requerimento')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = PedidoRequerimentoForm(instance=tipo)
    
    return render(request, 'precapp/tipo_pedido_requerimento_form.html', {
        'form': form,
        'tipo': tipo,
        'title': f'Editar Tipo de Pedido: {tipo.nome}'
    })


@login_required
def deletar_tipo_pedido_requerimento_view(request, tipo_id):
    """Delete a tipo de pedido de requerimento"""
    tipo = get_object_or_404(PedidoRequerimento, id=tipo_id)
    
    # Check if tipo is being used
    requerimentos_count = tipo.requerimento_set.count()
    
    if requerimentos_count > 0:
        messages.error(
            request, 
            f'Não é possível excluir o tipo de pedido "{tipo.nome}" pois ele está sendo usado por {requerimentos_count} requerimento(s).'
        )
        return redirect('tipos_pedido_requerimento')
    
    if request.method == 'POST':
        tipo_nome = tipo.nome
        tipo.delete()
        messages.success(request, f'Tipo de Pedido "{tipo_nome}" excluído com sucesso!')
        return redirect('tipos_pedido_requerimento')
    
    return render(request, 'precapp/confirmar_delete_tipo_pedido_requerimento.html', {
        'tipo_pedido': tipo,
        'requerimentos_count': requerimentos_count
    })


@login_required
def ativar_tipo_pedido_requerimento_view(request, tipo_id):
    """Toggle activation status of a tipo de pedido de requerimento"""
    tipo = get_object_or_404(PedidoRequerimento, id=tipo_id)
    
    # Handle POST data (from AJAX calls)
    if request.method == 'POST':
        ativo_value = request.POST.get('ativo', '').strip()
        tipo.ativo = ativo_value.lower() == 'true'
    else:
        # Handle GET for backward compatibility
        ativo_param = request.GET.get('ativo', '').lower()
        if ativo_param == 'true':
            tipo.ativo = True
        elif ativo_param == 'false':
            tipo.ativo = False
        else:
            # Toggle if no parameter
            tipo.ativo = not tipo.ativo
    
    tipo.save()
    
    status = "ativado" if tipo.ativo else "desativado"
    messages.success(request, f'Tipo de Pedido de Requerimento "{tipo.nome}" {status} com sucesso!')
    
    return redirect('tipos_pedido_requerimento')


# ===============================
# CUSTOMIZAÇÃO VIEWS
# ===============================

@login_required
def customizacao_view(request):
    """Central customization page for managing phases, types, and diligence types"""
    # Get statistics for phases, types, and diligence types
    fases_principais = Fase.objects.all()
    fases_honorarios = FaseHonorariosContratuais.objects.all()
    fases_honorarios_sucumbenciais = FaseHonorariosSucumbenciais.objects.all()
    tipos_precatorio = Tipo.objects.all()
    tipos_diligencia = TipoDiligencia.objects.all()
    tipos_pedido_requerimento = PedidoRequerimento.objects.all()
    contas_bancarias = ContaBancaria.objects.all()
    
    context = {
        # Fases Principais stats
        'total_fases_principais': fases_principais.count(),
        'fases_principais_ativas': fases_principais.filter(ativa=True).count(),
        'fases_principais_inativas': fases_principais.filter(ativa=False).count(),
        
        # Fases Honorários stats
        'total_fases_honorarios': fases_honorarios.count(),
        'fases_honorarios_ativas': fases_honorarios.filter(ativa=True).count(),
        'fases_honorarios_inativas': fases_honorarios.filter(ativa=False).count(),
        
        # Fases Honorários Sucumbenciais stats
        'total_fases_honorarios_sucumbenciais': fases_honorarios_sucumbenciais.count(),
        'fases_honorarios_sucumbenciais_ativas': fases_honorarios_sucumbenciais.filter(ativa=True).count(),
        'fases_honorarios_sucumbenciais_inativas': fases_honorarios_sucumbenciais.filter(ativa=False).count(),
        
        # Tipos Precatório stats
        'total_tipos_precatorio': tipos_precatorio.count(),
        'tipos_precatorio_ativos': tipos_precatorio.filter(ativa=True).count(),
        'tipos_precatorio_inativos': tipos_precatorio.filter(ativa=False).count(),
        
        # Tipos Pedido Requerimento stats
        'total_tipos_pedido_requerimento': tipos_pedido_requerimento.count(),
        'tipos_pedido_requerimento_ativos': tipos_pedido_requerimento.filter(ativo=True).count(),
        'tipos_pedido_requerimento_inativos': tipos_pedido_requerimento.filter(ativo=False).count(),
        
        # Tipos Diligência stats
        'total_tipos_diligencia': tipos_diligencia.count(),
        'tipos_diligencia_ativos': tipos_diligencia.filter(ativo=True).count(),
        'tipos_diligencia_inativos': tipos_diligencia.filter(ativo=False).count(),
        
        # ContaBancaria stats
        'total_contas_bancarias': contas_bancarias.count(),
        
        # Recent items (last 5 of each type)
        'recent_fases_principais': fases_principais.order_by('-criado_em')[:5],
        'recent_fases_honorarios': fases_honorarios.order_by('-criado_em')[:5],
        'recent_fases_honorarios_sucumbenciais': fases_honorarios_sucumbenciais.order_by('-criado_em')[:5],
        'recent_tipos_precatorio': tipos_precatorio.order_by('-criado_em')[:5],
        'recent_tipos_pedido_requerimento': tipos_pedido_requerimento.order_by('-criado_em')[:5],
        'recent_tipos_diligencia': tipos_diligencia.order_by('-criado_em')[:5],
        'recent_contas_bancarias': contas_bancarias.order_by('-criado_em')[:5],
    }
    
    return render(request, 'precapp/customizacao.html', context)


# ===============================
# DILIGENCIA VIEWS
# ===============================

@login_required
def nova_diligencia_view(request, cpf):
    """Create a new diligencia for a specific client"""
    cliente = get_object_or_404(Cliente, cpf=cpf)
    
    if request.method == 'POST':
        form = DiligenciasForm(request.POST)
        if form.is_valid():
            diligencia = form.save(commit=False)
            diligencia.cliente = cliente
            # Automatically set the creator to the current logged-in user
            diligencia.criado_por = request.user.get_full_name() or request.user.username
            diligencia.save()
            messages.success(request, f'Diligência criada com sucesso para {cliente.nome}!')
            return redirect('cliente_detail', cpf=cliente.cpf)
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = DiligenciasForm()
    
    context = {
        'form': form,
        'cliente': cliente,
        'action': 'create'
    }
    
    return render(request, 'precapp/diligencia_form.html', context)


@login_required
def editar_diligencia_view(request, cpf, diligencia_id):
    """Edit an existing diligencia"""
    cliente = get_object_or_404(Cliente, cpf=cpf)
    diligencia = get_object_or_404(Diligencias, id=diligencia_id, cliente=cliente)
    
    if request.method == 'POST':
        form = DiligenciasForm(request.POST, instance=diligencia)
        if form.is_valid():
            # Save the form but preserve the original creator
            updated_diligencia = form.save(commit=False)
            # Ensure the creator is not changed during edit
            updated_diligencia.criado_por = diligencia.criado_por
            updated_diligencia.save()
            messages.success(request, f'Diligência atualizada com sucesso!')
            return redirect('cliente_detail', cpf=cliente.cpf)
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = DiligenciasForm(instance=diligencia)
    
    context = {
        'form': form,
        'cliente': cliente,
        'diligencia': diligencia,
        'action': 'edit'
    }
    
    return render(request, 'precapp/diligencia_form.html', context)


@login_required
def deletar_diligencia_view(request, cpf, diligencia_id):
    """Delete a diligencia"""
    cliente = get_object_or_404(Cliente, cpf=cpf)
    diligencia = get_object_or_404(Diligencias, id=diligencia_id, cliente=cliente)
    
    if request.method == 'POST':
        diligencia_nome = str(diligencia)
        diligencia.delete()
        messages.success(request, f'Diligência "{diligencia_nome}" excluída com sucesso!')
        return redirect('cliente_detail', cpf=cliente.cpf)
    
    context = {
        'diligencia': diligencia,
        'cliente': cliente,
        'total_diligencias': cliente.diligencias.count(),
        'pendentes_diligencias': cliente.diligencias.filter(concluida=False).count(),
        'concluidas_diligencias': cliente.diligencias.filter(concluida=True).count(),
    }
    
    return render(request, 'precapp/confirmar_delete_diligencia.html', context)


@login_required
def marcar_diligencia_concluida_view(request, cpf, diligencia_id):
    """Toggle diligencia completion status"""
    cliente = get_object_or_404(Cliente, cpf=cpf)
    diligencia = get_object_or_404(Diligencias, id=diligencia_id, cliente=cliente)
    
    if request.method == 'POST':
        # Save the original state from database before form processing
        # since ModelForm.is_valid() can modify the instance
        original_diligencia = Diligencias.objects.get(id=diligencia.id)
        was_completed_before = original_diligencia.concluida
        
        form = DiligenciasUpdateForm(request.POST, instance=diligencia)
        if form.is_valid():
            updated_diligencia = form.save(commit=False)
            
            # Handle concluido_por field automatically
            if updated_diligencia.concluida:
                # If marking as completed, set the current user as completer
                if not was_completed_before:  # Use saved state from database
                    updated_diligencia.concluido_por = request.user.get_full_name() or request.user.username
            else:
                # If marking as not completed (reopening), clear the completer
                updated_diligencia.concluido_por = None
            
            updated_diligencia.save()
            status = "concluída" if updated_diligencia.concluida else "reaberta"
            messages.success(request, f'Diligência marcada como {status}!')
            return redirect('cliente_detail', cpf=cliente.cpf)
        else:
            messages.error(request, 'Erro ao atualizar a diligência.')
    else:
        form = DiligenciasUpdateForm(instance=diligencia)
    
    # Calculate statistics for the client
    total_diligencias = cliente.diligencias.count()
    pendentes_diligencias = cliente.diligencias.filter(concluida=False).count()
    concluidas_diligencias = cliente.diligencias.filter(concluida=True).count()
    
    context = {
        'form': form,
        'diligencia': diligencia,
        'cliente': cliente,
        'total_diligencias': total_diligencias,
        'pendentes_diligencias': pendentes_diligencias,
        'concluidas_diligencias': concluidas_diligencias,
    }
    
    return render(request, 'precapp/diligencia_conclusao_form.html', context)


@login_required
def diligencias_list_view(request):
    """List all diligencias with filtering and search capabilities"""
    diligencias = Diligencias.objects.select_related('cliente', 'tipo', 'responsavel').order_by('-data_criacao')
    
    # Filter parameters
    status_filter = request.GET.get('status', '')
    urgencia_filter = request.GET.get('urgencia', '')
    tipo_filter = request.GET.get('tipo', '')
    responsavel_filter = request.GET.get('responsavel', '')
    search_query = request.GET.get('search', '')
    data_inicio_filter = request.GET.get('data_inicio', '')
    data_fim_filter = request.GET.get('data_fim', '')
    
    # Apply filters
    if status_filter == 'pendente':
        diligencias = diligencias.filter(concluida=False)
    elif status_filter == 'concluida':
        diligencias = diligencias.filter(concluida=True)
    
    if urgencia_filter:
        diligencias = diligencias.filter(urgencia=urgencia_filter)
    
    if tipo_filter:
        diligencias = diligencias.filter(tipo_id=tipo_filter)
    
    if responsavel_filter:
        diligencias = diligencias.filter(responsavel_id=responsavel_filter)
    
    if search_query:
        diligencias = diligencias.filter(
            Q(cliente__nome__icontains=search_query) |
            Q(cliente__cpf__icontains=search_query) |
            Q(tipo__nome__icontains=search_query) |
            Q(descricao__icontains=search_query) |
            Q(responsavel__username__icontains=search_query) |
            Q(responsavel__first_name__icontains=search_query) |
            Q(responsavel__last_name__icontains=search_query)
        )
    
    # Date range filter for data_final (due date)
    if data_inicio_filter:
        from datetime import datetime
        try:
            data_inicio = datetime.strptime(data_inicio_filter, '%Y-%m-%d').date()
            diligencias = diligencias.filter(data_final__gte=data_inicio)
        except ValueError:
            pass  # Invalid date format, ignore filter
    
    if data_fim_filter:
        from datetime import datetime
        try:
            data_fim = datetime.strptime(data_fim_filter, '%Y-%m-%d').date()
            diligencias = diligencias.filter(data_final__lte=data_fim)
        except ValueError:
            pass  # Invalid date format, ignore filter
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(diligencias, 25)  # 25 diligencias per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_diligencias = Diligencias.objects.count()
    pendentes_diligencias = Diligencias.objects.filter(concluida=False).count()
    concluidas_diligencias = Diligencias.objects.filter(concluida=True).count()
    atrasadas_diligencias = Diligencias.objects.filter(
        concluida=False,
        data_final__lt=timezone.now().date()
    ).count()
    
    # Get filter options
    tipos_diligencia = TipoDiligencia.get_ativos()
    urgencia_choices = Diligencias.URGENCIA_CHOICES
    
    # Get users who are assigned as responsavel (active users who have diligencias assigned)
    from django.contrib.auth.models import User
    usuarios_responsaveis = User.objects.filter(
        is_active=True,
        diligencias_responsavel__isnull=False
    ).distinct().order_by('first_name', 'last_name', 'username')
    
    context = {
        'page_obj': page_obj,
        'diligencias': page_obj,
        'total_diligencias': total_diligencias,
        'pendentes_diligencias': pendentes_diligencias,
        'concluidas_diligencias': concluidas_diligencias,
        'atrasadas_diligencias': atrasadas_diligencias,
        'tipos_diligencia': tipos_diligencia,
        'urgencia_choices': urgencia_choices,
        'usuarios_responsaveis': usuarios_responsaveis,
        'status_filter': status_filter,
        'urgencia_filter': urgencia_filter,
        'tipo_filter': tipo_filter,
        'responsavel_filter': responsavel_filter,
        'search_query': search_query,
        'data_inicio_filter': data_inicio_filter,
        'data_fim_filter': data_fim_filter,
    }
    
    return render(request, 'precapp/diligencias_list.html', context)


# ===============================
# EXCEL IMPORT VIEW
# ===============================

@login_required
def import_excel_view(request):
    """View to handle Excel file upload and import precatorios"""
    if request.method == 'POST':
        uploaded_file = request.FILES.get('excel_file')
        
        if not uploaded_file:
            messages.error(request, 'Por favor, selecione um arquivo Excel para importar.')
            return redirect('precatorios')
        
        # Validate file extension
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Por favor, selecione um arquivo Excel válido (.xlsx ou .xls).')
            return redirect('precatorios')
        
        # Validate file size (limit to 10MB)
        if uploaded_file.size > 10 * 1024 * 1024:
            messages.error(request, 'O arquivo é muito grande. O limite é de 10MB.')
            return redirect('precatorios')
        
        try:
            # Save uploaded file temporarily
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                # Write uploaded file content to temporary file
                for chunk in uploaded_file.chunks():
                    temp_file.write(chunk)
                temp_file_path = temp_file.name
            
            try:
                # Capture command output
                output = StringIO()
                
                # Call the import_excel management command
                call_command('import_excel', '--file', temp_file_path, stdout=output, stderr=output)
                
                # Get the output
                command_output = output.getvalue()
                
                # Parse the output to get statistics
                imported_stats = {
                    'precatorios': 0,
                    'clientes': 0,
                    'requerimentos': 0
                }
                
                # Extract statistics from output
                lines = command_output.split('\n')
                for line in lines:
                    if 'Precatorios:' in line:
                        try:
                            imported_stats['precatorios'] = int(line.split(':')[1].strip())
                        except:
                            pass
                    elif 'Clientes:' in line:
                        try:
                            imported_stats['clientes'] = int(line.split(':')[1].strip())
                        except:
                            pass
                    elif 'Requerimentos:' in line:
                        try:
                            imported_stats['requerimentos'] = int(line.split(':')[1].strip())
                        except:
                            pass
                
                # Success message with statistics
                total_imported = sum(imported_stats.values())
                if total_imported > 0:
                    stats_msg = []
                    if imported_stats['precatorios'] > 0:
                        stats_msg.append(f"{imported_stats['precatorios']} precatório(s)")
                    if imported_stats['clientes'] > 0:
                        stats_msg.append(f"{imported_stats['clientes']} cliente(s)")
                    if imported_stats['requerimentos'] > 0:
                        stats_msg.append(f"{imported_stats['requerimentos']} requerimento(s)")
                    
                    message = f"Importação concluída com sucesso! Importados: {', '.join(stats_msg)}."
                    messages.success(request, message)
                else:
                    messages.warning(request, 'Nenhum dado novo foi importado. Os dados podem já existir no sistema.')
                
            except Exception as e:
                # Error in command execution
                error_msg = str(e)
                if 'already exists' in error_msg.lower():
                    messages.warning(request, 'Alguns dados já existem no sistema e foram ignorados.')
                else:
                    messages.error(request, f'Erro durante a importação: {error_msg}')
                    
            finally:
                # Clean up temporary file
                try:
                    os.unlink(temp_file_path)
                except:
                    pass
                    
        except Exception as e:
            messages.error(request, f'Erro ao processar o arquivo: {str(e)}')
        
        return redirect('precatorios')
    
    # For GET requests, redirect to precatorios list
    return redirect('precatorios')


# ===============================
# EXCEL EXPORT FUNCTIONALITY
# ===============================

@login_required
def export_precatorios_excel(request):
    """
    Export comprehensive precatorios and related data to Excel format.
    
    This view is restricted to superusers only for security purposes.
    
    Creates a detailed Excel report with multiple sheets containing:
    - Complete precatorios data with client information
    - Client summary with associated precatorios count
    - Diligencias (legal tasks) data with status tracking
    - Requerimentos (legal requests) with financial data
    - Alvarás (payment authorizations) with multiple fee types
    - Statistical summary with comprehensive metrics
    
    Features:
    - Multiple worksheets for organized data presentation
    - Professional formatting with headers and color coding
    - Conditional formatting for status visualization
    - Brazilian date and currency formatting
    - Comprehensive relationship data
    - Business intelligence analytics
    
    Worksheets included:
    1. Precatórios: Main documents with client associations
    2. Clientes: Client summaries with aggregated data
    3. Diligências: Legal tasks with completion tracking
    4. Requerimentos: Legal requests with financial analysis
    5. Alvarás: Payment authorizations with fee breakdown
    6. Estatísticas: System-wide statistics and report metadata
    
    Returns:
        HttpResponse: Excel file download response with comprehensive data
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils import timezone
    import io
    
    # Check if user is superuser
    if not request.user.is_superuser:
        messages.error(request, 'Acesso negado. Apenas superusuários podem exportar dados.')
        return redirect('precatorio_view')
    
    # Create workbook
    wb = Workbook()
    
    # ==================== PRECATORIOS SHEET ====================
    ws_precatorios = wb.active
    ws_precatorios.title = "Precatórios"
    
    # Headers for precatorios sheet
    precatorio_headers = [
        'CNJ', 'Origem', 'Valor de Face', 'Última Atualização', 'Data Última Atualização',
        'Crédito Principal', 'Honorários Contratuais', 'Honorários Sucumbenciais',
        'Cliente Nome', 'Cliente CPF', 'Cliente Nascimento', 'Cliente Prioritário',
        'Tipo Precatório', 'Orçamento'
    ]
    
    # Write headers
    for col, header in enumerate(precatorio_headers, 1):
        cell = ws_precatorios.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get precatorios data with related information
    precatorios = Precatorio.objects.prefetch_related('clientes').select_related('tipo').all().order_by('cnj')
    
    # Write precatorios data
    for row, precatorio in enumerate(precatorios, 2):
        # Get the first client (since it's many-to-many, we'll take the first one for display)
        primeiro_cliente = precatorio.clientes.first()
        
        data = [
            precatorio.cnj,
            precatorio.origem,
            precatorio.valor_de_face,
            precatorio.ultima_atualizacao,
            precatorio.data_ultima_atualizacao.strftime('%d/%m/%Y') if precatorio.data_ultima_atualizacao else '',
            precatorio.get_credito_principal_display(),
            precatorio.get_honorarios_contratuais_display(),
            precatorio.get_honorarios_sucumbenciais_display(),
            primeiro_cliente.nome if primeiro_cliente else 'Não vinculado',
            primeiro_cliente.cpf if primeiro_cliente else '',
            primeiro_cliente.nascimento.strftime('%d/%m/%Y') if primeiro_cliente and primeiro_cliente.nascimento else '',
            'Sim' if primeiro_cliente and primeiro_cliente.prioridade else 'Não',
            precatorio.tipo.nome if precatorio.tipo else 'Não especificado',
            precatorio.orcamento
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_precatorios.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Format currency columns
            if col in [3, 4]:  # Valor de Face, Última Atualização
                if value is not None:
                    cell.number_format = 'R$ #,##0.00'
    
    # Auto-adjust column widths
    for col in range(1, len(precatorio_headers) + 1):
        ws_precatorios.column_dimensions[get_column_letter(col)].width = 15
    
    # ==================== CLIENTES SHEET ====================
    ws_clientes = wb.create_sheet(title="Clientes")
    
    # Headers for clientes sheet
    cliente_headers = [
        'Nome', 'CPF', 'Data Nascimento', 'Prioritário',
        'Total Precatórios', 'Valor Total Precatórios', 'Total Diligências',
        'Diligências Pendentes', 'Diligências Concluídas'
    ]
    
    # Write headers
    for col, header in enumerate(cliente_headers, 1):
        cell = ws_clientes.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get clientes data with aggregated information
    clientes = Cliente.objects.prefetch_related('precatorios', 'diligencias').all().order_by('nome')
    
    # Write clientes data
    for row, cliente in enumerate(clientes, 2):
        # Calculate aggregated data
        precatorios_count = cliente.precatorios.count()
        total_valor = sum(p.ultima_atualizacao or 0 for p in cliente.precatorios.all())
        diligencias_count = cliente.diligencias.count()
        diligencias_pendentes = cliente.diligencias.filter(concluida=False).count()
        diligencias_concluidas = cliente.diligencias.filter(concluida=True).count()
        
        data = [
            cliente.nome,
            cliente.cpf,
            cliente.nascimento.strftime('%d/%m/%Y') if cliente.nascimento else '',
            'Sim' if cliente.prioridade else 'Não',
            precatorios_count,
            total_valor,
            diligencias_count,
            diligencias_pendentes,
            diligencias_concluidas
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_clientes.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Format currency column
            if col == 6:  # Valor Total Precatórios
                cell.number_format = 'R$ #,##0.00'
            
            # Color code priority clients
            if col == 4 and value == 'Sim':
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(cliente_headers) + 1):
        ws_clientes.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== DILIGENCIAS SHEET ====================
    ws_diligencias = wb.create_sheet(title="Diligências")
    
    # Headers for diligencias sheet
    diligencia_headers = [
        'Cliente Nome', 'Cliente CPF', 'Tipo Diligência', 'Descrição',
        'Data Final', 'Urgência', 'Status', 'Data Conclusão',
        'Responsável', 'Criado Por', 'Concluído Por', 'Data Criação'
    ]
    
    # Write headers
    for col, header in enumerate(diligencia_headers, 1):
        cell = ws_diligencias.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get diligencias data with related information
    diligencias = Diligencias.objects.select_related(
        'cliente', 'tipo', 'responsavel'
    ).all().order_by('cliente__nome', 'data_final')
    
    # Write diligencias data
    for row, diligencia in enumerate(diligencias, 2):
        data = [
            diligencia.cliente.nome,
            diligencia.cliente.cpf,
            diligencia.tipo.nome,
            diligencia.descricao[:100] + '...' if len(diligencia.descricao or '') > 100 else diligencia.descricao,
            diligencia.data_final.strftime('%d/%m/%Y') if diligencia.data_final else '',
            diligencia.get_urgencia_display(),
            'Concluída' if diligencia.concluida else 'Pendente',
            diligencia.data_conclusao.strftime('%d/%m/%Y %H:%M') if diligencia.data_conclusao else '',
            diligencia.responsavel.get_full_name() if diligencia.responsavel else 'Não atribuído',
            diligencia.criado_por or '',
            diligencia.concluido_por or '',
            diligencia.criado_em.strftime('%d/%m/%Y %H:%M') if diligencia.criado_em else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_diligencias.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Color code status
            if col == 7:  # Status column
                if value == 'Concluída':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif value == 'Pendente':
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # Color code urgency
            if col == 6:  # Urgência column
                if value == 'Alta':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                elif value == 'Média':
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(diligencia_headers) + 1):
        ws_diligencias.column_dimensions[get_column_letter(col)].width = 20
    
    # ==================== REQUERIMENTOS SHEET ====================
    ws_requerimentos = wb.create_sheet(title="Requerimentos")
    
    # Headers for requerimentos sheet
    requerimento_headers = [
        'Cliente Nome', 'Cliente CPF', 'Precatório CNJ', 'Tipo Pedido',
        'Valor', 'Deságio (%)', 'Fase Atual', 'Data Criação',
        'Status da Fase', 'Valor com Deságio'
    ]
    
    # Write headers
    for col, header in enumerate(requerimento_headers, 1):
        cell = ws_requerimentos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get requerimentos data with related information
    requerimentos = Requerimento.objects.select_related(
        'cliente', 'precatorio', 'pedido', 'fase'
    ).all().order_by('cliente__nome', 'precatorio__cnj')
    
    # Write requerimentos data
    for row, requerimento in enumerate(requerimentos, 2):
        # Calculate valor com deságio - ensure safe calculation
        try:
            if requerimento.valor and requerimento.desagio:
                valor_com_desagio = requerimento.valor * (1 - requerimento.desagio / 100)
            else:
                valor_com_desagio = requerimento.valor or 0
        except (TypeError, ZeroDivisionError):
            valor_com_desagio = requerimento.valor or 0
        
        data = [
            requerimento.cliente.nome if requerimento.cliente else 'Não vinculado',
            requerimento.cliente.cpf if requerimento.cliente else '',
            requerimento.precatorio.cnj if requerimento.precatorio else '',
            requerimento.pedido.nome if requerimento.pedido else 'Não especificado',
            requerimento.valor or 0,
            requerimento.desagio or 0,
            requerimento.fase.nome if requerimento.fase else 'Sem fase',
            requerimento.precatorio.data_ultima_atualizacao.strftime('%d/%m/%Y') if requerimento.precatorio and requerimento.precatorio.data_ultima_atualizacao else '',
            'Ativa' if requerimento.fase and requerimento.fase.ativa else 'Inativa' if requerimento.fase else 'N/A',
            valor_com_desagio
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_requerimentos.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Format currency columns
            if col in [5, 10]:  # Valor and Valor com Deságio
                if value is not None and value != 0:
                    cell.number_format = 'R$ #,##0.00'
            
            # Format percentage column - simplified approach
            if col == 6:  # Deságio
                if value is not None and value != 0:
                    cell.number_format = '0.00'  # Just show as number with 2 decimals
                    # Add % symbol manually to the value display
                    cell.value = f"{value}%"
            
            # Color code status
            if col == 9:  # Status da Fase
                if value == 'Ativa':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif value == 'Inativa':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(requerimento_headers) + 1):
        ws_requerimentos.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== ALVARÁS SHEET ====================
    ws_alvaras = wb.create_sheet(title="Alvarás")
    
    # Headers for alvarás sheet
    alvara_headers = [
        'Cliente Nome', 'Cliente CPF', 'Precatório CNJ', 'Valor Principal',
        'Honorários Contratuais', 'Honorários Sucumbenciais', 'Valor Total',
        'Tipo Alvará', 'Fase Principal', 'Fase Honorários', 'Status Geral'
    ]
    
    # Write headers
    for col, header in enumerate(alvara_headers, 1):
        cell = ws_alvaras.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get alvarás data with related information
    alvaras = Alvara.objects.select_related(
        'cliente', 'precatorio', 'fase', 'fase_honorarios_contratuais'
    ).all().order_by('cliente__nome', 'precatorio__cnj')
    
    # Write alvarás data
    for row, alvara in enumerate(alvaras, 2):
        # Calculate total value
        valor_total = (alvara.valor_principal or 0) + (alvara.honorarios_contratuais or 0) + (alvara.honorarios_sucumbenciais or 0)
        
        # Determine overall status
        status_geral = 'N/A'
        if alvara.fase and alvara.fase_honorarios_contratuais:
            if alvara.fase.ativa and alvara.fase_honorarios_contratuais.ativa:
                status_geral = 'Ambas Ativas'
            elif not alvara.fase.ativa and not alvara.fase_honorarios_contratuais.ativa:
                status_geral = 'Ambas Inativas'
            else:
                status_geral = 'Misto'
        elif alvara.fase:
            status_geral = 'Ativa' if alvara.fase.ativa else 'Inativa'
        elif alvara.fase_honorarios_contratuais:
            status_geral = 'Hon. Ativa' if alvara.fase_honorarios_contratuais.ativa else 'Hon. Inativa'
        
        data = [
            alvara.cliente.nome,
            alvara.cliente.cpf,
            alvara.precatorio.cnj,
            alvara.valor_principal,
            alvara.honorarios_contratuais,
            alvara.honorarios_sucumbenciais,
            valor_total,
            alvara.tipo,
            alvara.fase.nome if alvara.fase else 'Sem fase',
            alvara.fase_honorarios_contratuais.nome if alvara.fase_honorarios_contratuais else 'Sem fase',
            status_geral
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_alvaras.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Format currency columns
            if col in [4, 5, 6, 7]:  # All monetary values
                if value is not None:
                    cell.number_format = 'R$ #,##0.00'
            
            # Color code status
            if col == 11:  # Status Geral
                if 'Ativa' in value and 'Inativa' not in value:
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif 'Inativa' in value and 'Ativa' not in value:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                elif 'Misto' in value:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(alvara_headers) + 1):
        ws_alvaras.column_dimensions[get_column_letter(col)].width = 18

    # ==================== RECEBIMENTOS SHEET ====================
    ws_recebimentos = wb.create_sheet(title="Recebimentos")
    
    # Headers for recebimentos sheet
    recebimento_headers = [
        'Número Documento', 'Alvará ID', 'Data', 'Conta Bancária', 
        'Valor', 'Tipo', 'Cliente Nome', 'Cliente CPF',
        'Precatório CNJ', 'Criado Em', 'Criado Por'
    ]
    
    # Write headers
    for col, header in enumerate(recebimento_headers, 1):
        cell = ws_recebimentos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="795548", end_color="795548", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get recebimentos data with related information
    recebimentos = Recebimentos.objects.select_related(
        'alvara', 'alvara__cliente', 'alvara__precatorio', 'conta_bancaria'
    ).all().order_by('-data', 'numero_documento')
    
    # Write recebimentos data
    for row, recebimento in enumerate(recebimentos, 2):
        data = [
            recebimento.numero_documento,
            f"ALV-{recebimento.alvara.id}" if recebimento.alvara else '',
            recebimento.data.strftime('%d/%m/%Y') if recebimento.data else '',
            f"{recebimento.conta_bancaria.banco} - {recebimento.conta_bancaria.agencia}/{recebimento.conta_bancaria.conta}" if recebimento.conta_bancaria else '',
            float(recebimento.valor) if recebimento.valor else 0,
            recebimento.get_tipo_display(),
            recebimento.alvara.cliente.nome if recebimento.alvara and recebimento.alvara.cliente else '',
            recebimento.alvara.cliente.cpf if recebimento.alvara and recebimento.alvara.cliente else '',
            recebimento.alvara.precatorio.cnj if recebimento.alvara and recebimento.alvara.precatorio else '',
            recebimento.criado_em.strftime('%d/%m/%Y %H:%M') if recebimento.criado_em else '',
            recebimento.criado_por or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_recebimentos.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Format currency column
            if col == 5:  # Valor
                if value is not None and value != 0:
                    cell.number_format = 'R$ #,##0.00'
            
            # Color code by tipo
            if col == 6:  # Tipo
                if 'contratuais' in str(value).lower():
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                elif 'sucumbenciais' in str(value).lower():
                    cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(recebimento_headers) + 1):
        ws_recebimentos.column_dimensions[get_column_letter(col)].width = 15

    # ==================== STATISTICS SHEET ====================
    ws_stats = wb.create_sheet(title="Estatísticas")
    
    # Calculate statistics
    total_precatorios = Precatorio.objects.count()
    total_clientes = Cliente.objects.count()
    total_diligencias = Diligencias.objects.count()
    total_requerimentos = Requerimento.objects.count()
    total_alvaras = Alvara.objects.count()
    total_recebimentos = Recebimentos.objects.count()
    clientes_prioritarios = Cliente.objects.filter(prioridade=True).count()
    diligencias_pendentes = Diligencias.objects.filter(concluida=False).count()
    diligencias_concluidas = Diligencias.objects.filter(concluida=True).count()
    valor_total_precatorios = sum(p.ultima_atualizacao or 0 for p in Precatorio.objects.all())
    valor_total_requerimentos = sum(r.valor or 0 for r in Requerimento.objects.all())
    valor_total_alvaras = sum((a.valor_principal or 0) + (a.honorarios_contratuais or 0) + (a.honorarios_sucumbenciais or 0) for a in Alvara.objects.all())
    valor_total_recebimentos = sum(r.valor or 0 for r in Recebimentos.objects.all())
    
    # Requerimentos by status
    requerimentos_com_fase = Requerimento.objects.exclude(fase__isnull=True).count()
    requerimentos_sem_fase = total_requerimentos - requerimentos_com_fase
    
    # Alvarás by type
    alvaras_aguardando_deposito = Alvara.objects.filter(tipo__icontains='aguardando').count()
    alvaras_deposito_judicial = Alvara.objects.filter(tipo__icontains='depósito').count()
    alvaras_recebido_cliente = Alvara.objects.filter(tipo__icontains='recebido').count()
    
    # Recebimentos by type
    recebimentos_contratuais = Recebimentos.objects.filter(tipo='Hon. contratuais').count()
    recebimentos_sucumbenciais = Recebimentos.objects.filter(tipo='Hon. sucumbenciais').count()
    valor_recebimentos_contratuais = sum(r.valor or 0 for r in Recebimentos.objects.filter(tipo='Hon. contratuais'))
    valor_recebimentos_sucumbenciais = sum(r.valor or 0 for r in Recebimentos.objects.filter(tipo='Hon. sucumbenciais'))
    
    # Statistics data
    stats_data = [
        ['Estatística', 'Valor'],
        ['', ''],
        ['### DOCUMENTOS ###', ''],
        ['Total de Precatórios', total_precatorios],
        ['Total de Requerimentos', total_requerimentos],
        ['Total de Alvarás', total_alvaras],
        ['Total de Recebimentos', total_recebimentos],
        ['', ''],
        ['### CLIENTES ###', ''],
        ['Total de Clientes', total_clientes],
        ['Clientes Prioritários', clientes_prioritarios],
        ['', ''],
        ['### DILIGÊNCIAS ###', ''],
        ['Total de Diligências', total_diligencias],
        ['Diligências Pendentes', diligencias_pendentes],
        ['Diligências Concluídas', diligencias_concluidas],
        ['', ''],
        ['### REQUERIMENTOS ###', ''],
        ['Requerimentos com Fase', requerimentos_com_fase],
        ['Requerimentos sem Fase', requerimentos_sem_fase],
        ['', ''],
        ['### ALVARÁS ###', ''],
        ['Alvarás - Aguardando Depósito', alvaras_aguardando_deposito],
        ['Alvarás - Depósito Judicial', alvaras_deposito_judicial],
        ['Alvarás - Recebido pelo Cliente', alvaras_recebido_cliente],
        ['', ''],
        ['### RECEBIMENTOS ###', ''],
        ['Recebimentos - Hon. Contratuais', recebimentos_contratuais],
        ['Recebimentos - Hon. Sucumbenciais', recebimentos_sucumbenciais],
        ['', ''],
        ['### VALORES FINANCEIROS ###', ''],
        ['Valor Total dos Precatórios', valor_total_precatorios],
        ['Valor Total dos Requerimentos', valor_total_requerimentos],
        ['Valor Total dos Alvarás', valor_total_alvaras],
        ['Valor Total dos Recebimentos', valor_total_recebimentos],
        ['Valor Recebimentos Contratuais', valor_recebimentos_contratuais],
        ['Valor Recebimentos Sucumbenciais', valor_recebimentos_sucumbenciais],
        ['', ''],
        ['### RELATÓRIO ###', ''],
        ['Data do Relatório', timezone.now().strftime('%d/%m/%Y %H:%M')],
        ['Gerado por', request.user.get_full_name() or request.user.username]
    ]
    
    # Write statistics
    for row, (label, value) in enumerate(stats_data, 1):
        # Ensure values are safe for Excel
        safe_label = str(label) if label is not None else ''
        safe_value = value if value is not None else ''
        
        # Label column
        cell_label = ws_stats.cell(row=row, column=1, value=safe_label)
        cell_label.font = Font(bold=True)
        cell_label.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Value column
        cell_value = ws_stats.cell(row=row, column=2, value=safe_value)
        cell_value.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format currency rows
        if 'Valor Total' in safe_label and isinstance(safe_value, (int, float)) and safe_value > 0:
            cell_value.number_format = 'R$ #,##0.00'
        
        # Header row formatting
        if row == 1:
            cell_label.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell_label.font = Font(bold=True, color="FFFFFF")
            cell_value.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell_value.font = Font(bold=True, color="FFFFFF")
        
        # Section header formatting (labels that start with ###)
        if isinstance(safe_label, str) and safe_label.startswith('###'):
            cell_label.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell_label.font = Font(bold=True, color="000000")
            cell_value.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Auto-adjust column widths
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20
    
    # ==================== PREPARE RESPONSE ====================
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with timestamp
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'relatorio_completo_sistema_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
def export_clientes_excel(request):
    """
    Export comprehensive clientes data to Excel format.
    
    This view is restricted to superusers only for security purposes.
    
    Creates a detailed Excel report focused on client information with:
    - Complete client data with associated precatorios
    - Client diligencias summary
    - Statistical analysis by priority status
    - Financial summary per client
    
    Features:
    - Professional formatting with headers and borders
    - Brazilian date and currency formatting
    - Color coding for priority clients
    - Comprehensive relationship data
    - Client-focused statistics
    
    Returns:
        HttpResponse: Excel file download response
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils import timezone
    from datetime import date, timedelta
    import io
    
    # Check if user is superuser
    if not request.user.is_superuser:
        messages.error(request, 'Acesso negado. Apenas superusuários podem exportar dados.')
        return redirect('clientes_view')
    
    # Create workbook
    wb = Workbook()
    
    # ==================== CLIENTES DETALHADO SHEET ====================
    ws_clientes = wb.active
    ws_clientes.title = "Clientes Detalhado"
    
    # Headers for detailed clients sheet
    cliente_headers = [
        'CPF/CNPJ', 'Nome Completo', 'Data Nascimento', 'Idade', 'Cliente Prioritário',
        'Total Precatórios', 'CNJ Precatórios', 'Valor Total Precatórios',
        'Total Diligências', 'Diligências Pendentes', 'Diligências Concluídas',
        'Diligências Atrasadas', 'Última Diligência', 'Próximo Vencimento'
    ]
    
    # Write headers
    for col, header in enumerate(cliente_headers, 1):
        cell = ws_clientes.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get clientes data with all related information
    clientes = Cliente.objects.prefetch_related(
        'precatorios', 'diligencias', 'diligencias__tipo'
    ).all().order_by('nome')
    
    # Write clientes data
    for row, cliente in enumerate(clientes, 2):
        from datetime import date, timedelta
        
        # Calculate client data
        precatorios_list = list(cliente.precatorios.all())
        precatorios_count = len(precatorios_list)
        cnj_list = ', '.join([p.cnj for p in precatorios_list[:3]])  # Show first 3 CNJs
        if len(precatorios_list) > 3:
            cnj_list += f' (e mais {len(precatorios_list) - 3})'
        
        total_valor = sum(p.ultima_atualizacao or p.valor_de_face or 0 for p in precatorios_list)
        
        # Diligencias calculations
        diligencias_all = list(cliente.diligencias.all())
        diligencias_count = len(diligencias_all)
        diligencias_pendentes = len([d for d in diligencias_all if not d.concluida])
        diligencias_concluidas = len([d for d in diligencias_all if d.concluida])
        
        # Calculate overdue diligencias
        today = date.today()
        diligencias_atrasadas = len([
            d for d in diligencias_all 
            if not d.concluida and d.data_final and d.data_final < today
        ])
        
        # Get last diligencia and next due date
        ultima_diligencia = ''
        proximo_vencimento = ''
        
        if diligencias_all:
            # Sort by creation date (most recent first)
            diligencias_ordenadas = sorted(
                diligencias_all, 
                key=lambda x: x.criado_em if hasattr(x, 'criado_em') and x.criado_em else date.min,
                reverse=True
            )
            if diligencias_ordenadas:
                ultima = diligencias_ordenadas[0]
                ultima_diligencia = f"{ultima.tipo.nome} - {ultima.descricao[:30]}..."
        
        # Find next due date from pending diligencias
        diligencias_pendentes_com_data = [
            d for d in diligencias_all 
            if not d.concluida and d.data_final and d.data_final >= today
        ]
        if diligencias_pendentes_com_data:
            proxima = min(diligencias_pendentes_com_data, key=lambda x: x.data_final)
            proximo_vencimento = proxima.data_final.strftime('%d/%m/%Y')
        
        # Calculate age if birth date is available
        idade = ''
        if cliente.nascimento:
            idade_anos = (date.today() - cliente.nascimento).days // 365
            idade = f"{idade_anos} anos"
        
        data = [
            cliente.cpf,
            cliente.nome,
            cliente.nascimento.strftime('%d/%m/%Y') if cliente.nascimento else '',
            idade,
            'Sim' if cliente.prioridade else 'Não',
            precatorios_count,
            cnj_list if cnj_list else 'Nenhum',
            total_valor,
            diligencias_count,
            diligencias_pendentes,
            diligencias_concluidas,
            diligencias_atrasadas,
            ultima_diligencia if ultima_diligencia else 'Nenhuma',
            proximo_vencimento if proximo_vencimento else 'Nenhum'
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_clientes.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Format currency column
            if col == 8:  # Valor Total Precatórios
                if isinstance(value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'
            
            # Color code priority clients (entire row)
            if col == 5 and value == 'Sim':  # Priority column
                for c in range(1, len(cliente_headers) + 1):
                    ws_clientes.cell(row=row, column=c).fill = PatternFill(
                        start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                    )
            
            # Color code overdue diligencias
            if col == 12 and isinstance(value, int) and value > 0:  # Diligências Atrasadas
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(cliente_headers) + 1):
        ws_clientes.column_dimensions[get_column_letter(col)].width = 20
    
    # ==================== RESUMO POR PRIORIDADE SHEET ====================
    ws_resumo = wb.create_sheet(title="Resumo por Prioridade")
    
    # Calculate priority statistics
    clientes_prioritarios = clientes.filter(prioridade=True)
    clientes_normais = clientes.filter(prioridade=False)
    
    total_clientes = clientes.count()
    total_prioritarios = clientes_prioritarios.count()
    total_normais = clientes_normais.count()
    
    # Calculate financial data
    valor_prioritarios = sum(
        sum(p.ultima_atualizacao or p.valor_de_face or 0 for p in c.precatorios.all())
        for c in clientes_prioritarios
    )
    valor_normais = sum(
        sum(p.ultima_atualizacao or p.valor_de_face or 0 for p in c.precatorios.all())
        for c in clientes_normais
    )
    
    # Summary data
    resumo_data = [
        ['Categoria', 'Quantidade Clientes', 'Percentual', 'Valor Total Precatórios', 'Valor Médio por Cliente'],
        ['Clientes Prioritários', total_prioritarios, f"{(total_prioritarios/total_clientes*100):.1f}%" if total_clientes > 0 else "0%", valor_prioritarios, valor_prioritarios/total_prioritarios if total_prioritarios > 0 else 0],
        ['Clientes Normais', total_normais, f"{(total_normais/total_clientes*100):.1f}%" if total_clientes > 0 else "0%", valor_normais, valor_normais/total_normais if total_normais > 0 else 0],
        ['', '', '', '', ''],
        ['TOTAL GERAL', total_clientes, '100%', valor_prioritarios + valor_normais, (valor_prioritarios + valor_normais)/total_clientes if total_clientes > 0 else 0]
    ]
    
    # Write summary data
    for row, data_row in enumerate(resumo_data, 1):
        for col, value in enumerate(data_row, 1):
            cell = ws_resumo.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Header row formatting
            if row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Total row formatting
            elif row == 5:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            
            # Priority clients row
            elif row == 2:
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # Format currency columns
            if col in [4, 5] and row > 1 and isinstance(value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
    
    # Auto-adjust column widths
    for col in range(1, 6):
        ws_resumo.column_dimensions[get_column_letter(col)].width = 25
    
    # ==================== ESTATÍSTICAS DETALHADAS SHEET ====================
    ws_stats = wb.create_sheet(title="Estatísticas Detalhadas")
    
    # Calculate detailed statistics
    total_diligencias_sistema = Diligencias.objects.count()
    diligencias_pendentes_sistema = Diligencias.objects.filter(concluida=False).count()
    diligencias_concluidas_sistema = Diligencias.objects.filter(concluida=True).count()
    diligencias_atrasadas_sistema = Diligencias.objects.filter(
        concluida=False, data_final__lt=date.today()
    ).count()
    
    total_precatorios_sistema = Precatorio.objects.count()
    valor_total_sistema = sum(p.ultima_atualizacao or p.valor_de_face or 0 for p in Precatorio.objects.all())
    
    # Statistics data
    stats_data = [
        ['Estatística', 'Valor'],
        ['### CLIENTES ###', ''],
        ['Total de Clientes', total_clientes],
        ['Clientes Prioritários', total_prioritarios],
        ['Clientes Normais', total_normais],
        ['Percentual Prioritários', f"{(total_prioritarios/total_clientes*100):.1f}%" if total_clientes > 0 else "0%"],
        ['', ''],
        ['### PRECATÓRIOS ###', ''],
        ['Total de Precatórios no Sistema', total_precatorios_sistema],
        ['Valor Total dos Precatórios', valor_total_sistema],
        ['Valor Médio por Precatório', valor_total_sistema/total_precatorios_sistema if total_precatorios_sistema > 0 else 0],
        ['', ''],
        ['### DILIGÊNCIAS ###', ''],
        ['Total de Diligências no Sistema', total_diligencias_sistema],
        ['Diligências Pendentes', diligencias_pendentes_sistema],
        ['Diligências Concluídas', diligencias_concluidas_sistema],
        ['Diligências Atrasadas', diligencias_atrasadas_sistema],
        ['Taxa de Conclusão', f"{(diligencias_concluidas_sistema/total_diligencias_sistema*100):.1f}%" if total_diligencias_sistema > 0 else "0%"],
        ['', ''],
        ['### RELATÓRIO ###', ''],
        ['Data do Relatório', timezone.now().strftime('%d/%m/%Y %H:%M')],
        ['Gerado por', request.user.get_full_name() or request.user.username],
        ['Tipo de Relatório', 'Exportação de Clientes']
    ]
    
    # Write statistics
    for row, (label, value) in enumerate(stats_data, 1):
        # Label column
        cell_label = ws_stats.cell(row=row, column=1, value=label)
        cell_label.font = Font(bold=True if '===' in str(label) else False)
        cell_label.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Value column
        cell_value = ws_stats.cell(row=row, column=2, value=value)
        cell_value.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format currency rows
        if 'Valor' in str(label) and 'Valor' != str(label) and isinstance(value, (int, float)):
            cell_value.number_format = 'R$ #,##0.00'
        
        # Header row formatting
        if row == 1:
            cell_label.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell_label.font = Font(bold=True, color="FFFFFF")
            cell_value.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell_value.font = Font(bold=True, color="FFFFFF")
        
        # Section headers
        elif '===' in str(label):
            cell_label.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            cell_value.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    
    # Auto-adjust column widths
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 25

    # ==================== DILIGENCIAS SHEET ====================
    ws_diligencias = wb.create_sheet(title="Diligências")
    
    # Headers for diligencias sheet
    diligencia_headers = [
        'Cliente Nome', 'Cliente CPF', 'Tipo Diligência', 'Descrição',
        'Data Final', 'Urgência', 'Status', 'Data Conclusão',
        'Responsável', 'Criado Por', 'Data Criação', 'Observações'
    ]
    
    # Write headers
    for col, header in enumerate(diligencia_headers, 1):
        cell = ws_diligencias.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Get all diligencias data with related information
    all_diligencias = Diligencias.objects.select_related(
        'cliente', 'tipo', 'responsavel'
    ).all().order_by('cliente__nome', 'data_final')
    
    # Write diligencias data
    for row, diligencia in enumerate(all_diligencias, 2):
        data = [
            diligencia.cliente.nome,
            diligencia.cliente.cpf,
            diligencia.tipo.nome,
            diligencia.descricao[:100] + '...' if diligencia.descricao and len(diligencia.descricao) > 100 else diligencia.descricao or '',
            diligencia.data_final.strftime('%d/%m/%Y') if diligencia.data_final else '',
            diligencia.get_urgencia_display() if hasattr(diligencia, 'urgencia') and diligencia.urgencia else 'Normal',
            'Concluída' if diligencia.concluida else 'Pendente',
            diligencia.data_conclusao.strftime('%d/%m/%Y %H:%M') if diligencia.data_conclusao else '',
            diligencia.responsavel.get_full_name() if diligencia.responsavel else '',
            diligencia.criado_por or '',
            diligencia.criado_em.strftime('%d/%m/%Y %H:%M') if hasattr(diligencia, 'criado_em') and diligencia.criado_em else '',
            diligencia.observacoes[:50] + '...' if hasattr(diligencia, 'observacoes') and diligencia.observacoes and len(diligencia.observacoes) > 50 else (diligencia.observacoes if hasattr(diligencia, 'observacoes') else '')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_diligencias.cell(row=row, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Color code by status
            if col == 7:  # Status column
                if value == 'Concluída':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif value == 'Pendente':
                    # Check if overdue
                    if diligencia.data_final and diligencia.data_final < date.today():
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # Color code priority clients
            if diligencia.cliente.prioridade and col == 1:  # Cliente Nome column
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(diligencia_headers) + 1):
        ws_diligencias.column_dimensions[get_column_letter(col)].width = 18

    # ==================== PREPARE RESPONSE ====================    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with timestamp
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'relatorio_clientes_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


# ===============================
# FILE DOWNLOAD VIEWS
# ===============================

@login_required
def download_precatorio_file(request, precatorio_cnj):
    """
    Download the integra_precatorio file for a specific precatorio
    Fixed to use direct file access instead of exists() check
    """
    from django.core.files.storage import default_storage
    from django.http import StreamingHttpResponse, Http404
    from django.utils.encoding import smart_str
    import mimetypes
    
    # Get precatorio object
    try:
        precatorio = get_object_or_404(Precatorio, cnj=precatorio_cnj)
        logger.info(f"Download request for precatorio: {precatorio_cnj}")
    except Exception as e:
        logger.error(f"Precatorio not found: {precatorio_cnj} - {e}")
        raise Http404("Precatório não encontrado")
    
    # Check if file exists
    if not precatorio.integra_precatorio:
        logger.error(f"No file for precatorio: {precatorio_cnj}")
        raise Http404("Nenhum arquivo encontrado para este precatório")
    
    file_name = precatorio.integra_precatorio.name
    logger.info(f"Attempting to download file: {file_name}")
    
    # Determine download filename
    if precatorio.integra_precatorio_filename:
        download_filename = precatorio.integra_precatorio_filename
        logger.info(f"Using stored original filename: {download_filename}")
    else:
        import os
        base_filename = os.path.basename(file_name)
        if base_filename and '.' in base_filename:
            download_filename = base_filename
            logger.info(f"Using filename from path: {download_filename}")
        else:
            download_filename = f"precatorio_{precatorio_cnj.replace('/', '_').replace('-', '_')}.pdf"
            logger.info(f"Using generated filename: {download_filename}")
    
    try:
        # Try to open the file directly (same as test download)
        file_obj = default_storage.open(file_name, 'rb')
        logger.info(f"Successfully opened file: {file_name}")
        
        # Get content type
        content_type, _ = mimetypes.guess_type(download_filename)
        if not content_type:
            content_type = 'application/pdf'  # Default to PDF
        
        # Create file iterator for streaming
        def file_iterator(file_obj, chunk_size=8192):
            try:
                while True:
                    chunk = file_obj.read(chunk_size)
                    if not chunk:
                        break
                    yield chunk
            finally:
                file_obj.close()
        
        # Create streaming response
        response = StreamingHttpResponse(
            file_iterator(file_obj),
            content_type=content_type
        )
        
        # Set download headers
        safe_filename = smart_str(download_filename.replace('"', ''))
        response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
        
        # Add file size if available
        try:
            file_size = default_storage.size(file_name)
            response['Content-Length'] = str(file_size)
            logger.info(f"Set Content-Length: {file_size}")
        except Exception as e:
            logger.warning(f"Could not get file size: {e}")
        
        # Prevent caching
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'

        logger.info(f"Returning streaming download response for: {download_filename}")
        return response

    except Exception as e:
        logger.error(f"Error accessing file {file_name}: {str(e)}")
        raise Http404(f"Erro ao acessar arquivo: {str(e)}")


# ==============================
# ContaBancaria CRUD Views
# ==============================

@login_required
def contas_bancarias_view(request):
    """View to list all bank accounts"""
    contas = ContaBancaria.objects.all()  # Uses model's default ordering: ['banco', 'agencia']
    
    # Add recebimentos count for each conta
    for conta in contas:
        conta.recebimentos_count = conta.recebimentos_set.count()
    
    # Statistics
    total_contas = contas.count()
    contas_com_recebimentos = sum(1 for conta in contas if conta.recebimentos_count > 0)
    contas_sem_recebimentos = total_contas - contas_com_recebimentos
    
    # Get count of total recebimentos using the reverse relationship
    total_recebimentos = sum(conta.recebimentos_count for conta in contas)
    
    context = {
        'contas_bancarias': contas,
        'total_contas': total_contas,
        'contas_com_recebimentos': contas_com_recebimentos,
        'contas_sem_recebimentos': contas_sem_recebimentos,
        'total_recebimentos': total_recebimentos,
    }
    
    return render(request, 'precapp/contas_bancarias_list.html', context)


@login_required
def nova_conta_bancaria_view(request):
    """View to create a new bank account"""
    if request.method == 'POST':
        form = ContaBancariaForm(request.POST)
        if form.is_valid():
            conta = form.save()
            messages.success(request, f'Conta Bancária "{conta.banco} - Ag: {conta.agencia}" criada com sucesso!')
            return redirect('contas_bancarias')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = ContaBancariaForm()
    
    context = {
        'form': form,
        'title': 'Nova Conta Bancária',
        'action': 'Criar'
    }
    
    return render(request, 'precapp/conta_bancaria_form.html', context)


@login_required
def editar_conta_bancaria_view(request, conta_id):
    """View to edit an existing bank account"""
    conta = get_object_or_404(ContaBancaria, id=conta_id)
    
    if request.method == 'POST':
        # Check if this is an update from the dropdown form in list view
        if 'update_conta' in request.POST:
            # Handle direct field updates from dropdown form
            banco = request.POST.get('banco')
            tipo_de_conta = request.POST.get('tipo_de_conta')
            agencia = request.POST.get('agencia')
            conta_numero = request.POST.get('conta')
            
            if banco and tipo_de_conta and agencia and conta_numero:
                conta.banco = banco
                conta.tipo_de_conta = tipo_de_conta
                conta.agencia = agencia
                conta.conta = conta_numero
                conta.save()
                messages.success(request, f'Conta Bancária "{conta.banco} - Ag: {conta.agencia}" atualizada com sucesso!')
                return redirect('contas_bancarias')
            else:
                messages.error(request, 'Todos os campos são obrigatórios.')
                return redirect('contas_bancarias')
        else:
            # Handle regular form update
            form = ContaBancariaForm(request.POST, instance=conta)
            if form.is_valid():
                conta = form.save()
                messages.success(request, f'Conta Bancária "{conta.banco} - Ag: {conta.agencia}" atualizada com sucesso!')
                return redirect('contas_bancarias')
            else:
                messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = ContaBancariaForm(instance=conta)
    
    context = {
        'form': form,
        'conta': conta,
        'title': f'Editar Conta: {conta.banco}',
        'action': 'Atualizar'
    }
    
    return render(request, 'precapp/conta_bancaria_form.html', context)


@login_required
def deletar_conta_bancaria_view(request, conta_id):
    """View to delete a bank account"""
    conta = get_object_or_404(ContaBancaria, id=conta_id)
    
    if request.method == 'POST':
        try:
            nome_conta = f"{conta.banco} - Ag: {conta.agencia}"
            conta.delete()
            messages.success(request, f'Conta Bancária "{nome_conta}" excluída com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao excluir conta: {str(e)}')
        
        return redirect('contas_bancarias')
    
    # Count related recebimentos
    recebimentos_count = conta.recebimentos_set.count()
    
    context = {
        'conta': conta,
        'recebimentos_count': recebimentos_count,
        'can_delete': recebimentos_count == 0
    }
    
    return render(request, 'precapp/confirmar_delete_conta_bancaria.html', context)


# ===============================
# RECEBIMENTOS VIEWS
# ===============================

@login_required
def novo_recebimento_view(request, alvara_id):
    """View to create a new payment for a specific alvará"""
    alvara = get_object_or_404(Alvara, id=alvara_id)
    
    if request.method == 'POST':
        form = RecebimentosForm(request.POST, alvara_id=alvara_id)
        if form.is_valid():
            recebimento = form.save(commit=False)
            recebimento.alvara = alvara
            recebimento.save()
            messages.success(request, f'Recebimento {recebimento.numero_documento} criado com sucesso!')
            return redirect('precatorio_detalhe', precatorio_cnj=alvara.precatorio.cnj)
        else:
            messages.error(request, 'Por favor, corrija os erros no formulário de recebimento.')
    else:
        form = RecebimentosForm(alvara_id=alvara_id)
    
    context = {
        'form': form,
        'alvara': alvara,
        'title': f'Novo Recebimento - Alvará {alvara.cliente.nome}',
        'action': 'Criar'
    }
    
    return render(request, 'precapp/recebimento_form.html', context)


@login_required
def listar_recebimentos_view(request, alvara_id):
    """View to list all payments for a specific alvará (AJAX)"""
    alvara = get_object_or_404(Alvara, id=alvara_id)
    recebimentos = alvara.recebimentos.all().order_by('-data', '-numero_documento')
    
    context = {
        'recebimentos': recebimentos,
        'alvara': alvara
    }
    
    return render(request, 'precapp/pagamentos_list_partial.html', context)


@login_required
def editar_recebimento_view(request, recebimento_id):
    """View to edit an existing payment"""
    recebimento = get_object_or_404(Recebimentos, numero_documento=recebimento_id)
    
    if request.method == 'POST':
        form = RecebimentosForm(request.POST, instance=recebimento, alvara_id=recebimento.alvara.id)
        if form.is_valid():
            recebimento = form.save()
            messages.success(request, f'Recebimento {recebimento.numero_documento} atualizado com sucesso!')
            return redirect('precatorio_detalhe', precatorio_cnj=recebimento.alvara.precatorio.cnj)
        else:
            messages.error(request, 'Por favor, corrija os erros no formulário.')
    else:
        form = RecebimentosForm(instance=recebimento, alvara_id=recebimento.alvara.id)
    
    context = {
        'form': form,
        'recebimento': recebimento,
        'alvara': recebimento.alvara,
        'title': f'Editar Recebimento {recebimento.numero_documento}',
        'action': 'Atualizar'
    }
    
    return render(request, 'precapp/recebimento_form.html', context)


@login_required
def deletar_recebimento_view(request, recebimento_id):
    """View to delete a receipt"""
    recebimento = get_object_or_404(Recebimentos, numero_documento=recebimento_id)
    
    if request.method == 'POST':
        alvara_precatorio_cnj = recebimento.alvara.precatorio.cnj
        numero_documento = recebimento.numero_documento
        try:
            recebimento.delete()
            messages.success(request, f'Recebimento {numero_documento} excluído com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao excluir recebimento: {str(e)}')
        
        return redirect('precatorio_detalhe', precatorio_cnj=alvara_precatorio_cnj)
    
    context = {
        'recebimento': recebimento,
        'alvara': recebimento.alvara
    }
    
    return render(request, 'precapp/confirmar_delete_recebimento.html', context)


# ===============================
# HELP VIEW
# ===============================

@login_required
def ajuda_view(request):
    """View to display help page with POP document link"""
    context = {
        'title': 'Ajuda - Sistema de Precatórios',
        'tipos_requerimento': PedidoRequerimento.get_ativos(),
    }
    
    return render(request, 'precapp/ajuda.html', context)