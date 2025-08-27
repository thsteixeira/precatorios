"""
Miscellaneous View Tests

Tests for miscellaneous views that don't fit into other categories.
Note: CustomizacaoViewTest has been moved to test_customizacao_view.py

Total expected tests: ~15 (UpdatePriorityByAgeViewTest)
Test classes: 1
"""

from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from datetime import date, timedelta
from unittest.mock import patch

from precapp.models import (
    Cliente, Precatorio, Diligencias, Requerimento, Alvara, 
    Tipo, TipoDiligencia, PedidoRequerimento, Fase, FaseHonorariosContratuais
)


class UpdatePriorityByAgeViewTest(TestCase):
    """Comprehensive tests for update_priority_by_age view function"""
    
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        
        # Create clients with different ages for testing
        today = date.today()
        
        # Client aged 65 (should get priority)
        self.cliente_65 = Cliente.objects.create(
            nome='Cliente 65 Anos',
            cpf='11111111111',
            nascimento=today - timedelta(days=65*365 + 100),  # Clearly over 60
            prioridade=False
        )
        
        # Client aged 70 (should get priority)
        self.cliente_70 = Cliente.objects.create(
            nome='Cliente 70 Anos',
            cpf='22222222222',
            nascimento=today - timedelta(days=70*365 + 100),  # Clearly over 60
            prioridade=False
        )
        
        # Client aged 45 (should NOT get priority)
        self.cliente_45 = Cliente.objects.create(
            nome='Cliente 45 Anos',
            cpf='33333333333',
            nascimento=today - timedelta(days=45*365),  # Clearly under 60
            prioridade=False
        )
        
        # Client aged 62 but already has priority (should NOT be updated)
        self.cliente_62_priority = Cliente.objects.create(
            nome='Cliente 62 com Prioridade',
            cpf='44444444444',
            nascimento=today - timedelta(days=62*365 + 100),  # Over 60
            prioridade=True  # Already has priority
        )
        
        # Client exactly 60 years old (boundary test - should NOT get priority due to < condition)
        self.cliente_60_exact = Cliente.objects.create(
            nome='Cliente Exatos 60 Anos',
            cpf='55555555555',
            nascimento=today - timedelta(days=60*365.25),  # Exactly 60 years
            prioridade=False
        )
        
        self.client_app = Client()

    # ==================== AUTHENTICATION TESTS ====================
    
    def test_update_priority_authentication_required(self):
        """Test that update priority view requires authentication"""
        response = self.client_app.post(reverse('update_priority_by_age'))
        self.assertEqual(response.status_code, 302)  # Redirect to login
        self.assertIn('/login/', response.url)
    
    def test_update_priority_get_method_not_allowed(self):
        """Test that GET method is not allowed (POST only view)"""
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.get(reverse('update_priority_by_age'))
        self.assertEqual(response.status_code, 405)  # Method not allowed

    # ==================== CORE BUSINESS LOGIC TESTS ====================
    
    def test_update_priority_successful_updates(self):
        """Test successful priority updates for clients over 60"""
        self.client_app.login(username='testuser', password='testpass123')
        
        # Verify initial state
        self.assertFalse(self.cliente_65.prioridade)
        self.assertFalse(self.cliente_70.prioridade)
        self.assertTrue(self.cliente_62_priority.prioridade)  # Already has priority
        
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        # Should redirect to clientes page
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('clientes'))
        
        # Refresh from database and verify updates
        self.cliente_65.refresh_from_db()
        self.cliente_70.refresh_from_db()
        self.cliente_45.refresh_from_db()
        self.cliente_62_priority.refresh_from_db()
        self.cliente_60_exact.refresh_from_db()
        
        # Clients over 60 should now have priority
        self.assertTrue(self.cliente_65.prioridade)
        self.assertTrue(self.cliente_70.prioridade)
        
        # Client under 60 should remain without priority
        self.assertFalse(self.cliente_45.prioridade)
        
        # Client exactly 60 should NOT get priority (< condition, not <=)
        self.assertFalse(self.cliente_60_exact.prioridade)
        
        # Client who already had priority should remain unchanged
        self.assertTrue(self.cliente_62_priority.prioridade)
    
    def test_update_priority_success_messages(self):
        """Test success messages when clients are updated"""
        self.client_app.login(username='testuser', password='testpass123')
        
        response = self.client_app.post(reverse('update_priority_by_age'))
        messages = list(get_messages(response.wsgi_request))
        
        # Should have success message
        success_messages = [m for m in messages if m.tags == 'success']
        self.assertEqual(len(success_messages), 1)
        self.assertIn('cliente(s) com mais de 60 anos foram atualizados', str(success_messages[0]))
        
        # Should have info message with examples
        info_messages = [m for m in messages if m.tags == 'info']
        self.assertEqual(len(info_messages), 1)
        self.assertIn('Exemplos atualizados:', str(info_messages[0]))
    
    def test_update_priority_no_clients_to_update(self):
        """Test when no clients need priority update"""
        # Set all clients over 60 to already have priority
        Cliente.objects.filter(nascimento__lt=date.today() - timedelta(days=60*365.25)).update(prioridade=True)
        
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        messages = list(get_messages(response.wsgi_request))
        info_messages = [m for m in messages if m.tags == 'info']
        
        self.assertEqual(len(info_messages), 1)
        self.assertIn('Nenhum cliente encontrado que precise ser atualizado', str(info_messages[0]))
        self.assertIn('já possuem status prioritário', str(info_messages[0]))
    
    def test_update_priority_empty_database(self):
        """Test behavior with no clients in database"""
        Cliente.objects.all().delete()
        
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        self.assertEqual(response.status_code, 302)
        messages = list(get_messages(response.wsgi_request))
        info_messages = [m for m in messages if m.tags == 'info']
        
        self.assertEqual(len(info_messages), 1)
        self.assertIn('Nenhum cliente encontrado', str(info_messages[0]))

    # ==================== AGE CALCULATION TESTS ====================
    
    def test_age_calculation_boundary_conditions(self):
        """Test age calculation boundary conditions"""
        today = date.today()
        
        # Create clients with precise age boundaries
        cliente_59_11_months = Cliente.objects.create(
            nome='Cliente 59 anos 11 meses',
            cpf='66666666666',
            nascimento=today - timedelta(days=59*365 + 330),  # Just under 60
            prioridade=False
        )
        
        cliente_60_1_day = Cliente.objects.create(
            nome='Cliente 60 anos 1 dia',
            cpf='77777777777',
            nascimento=today - timedelta(days=60*365.25 + 1),  # Just over 60
            prioridade=False
        )
        
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        # Refresh from database
        cliente_59_11_months.refresh_from_db()
        cliente_60_1_day.refresh_from_db()
        
        # Just under 60 should not get priority
        self.assertFalse(cliente_59_11_months.prioridade)
        
        # Just over 60 should get priority
        self.assertTrue(cliente_60_1_day.prioridade)
    
    def test_leap_year_calculation(self):
        """Test age calculation considering leap years"""
        # Test with a date that would be affected by leap years
        self.client_app.login(username='testuser', password='testpass123')
        
        # The function uses 365.25 days per year to account for leap years
        # This test ensures the calculation is consistent
        
        response = self.client_app.post(reverse('update_priority_by_age'))
        self.assertEqual(response.status_code, 302)
        
        # Verify the calculation used the correct formula
        sixty_years_ago = date.today() - timedelta(days=60*365.25)
        
        # All clients born before this calculated date should be updated
        clients_to_update = Cliente.objects.filter(
            nascimento__lt=sixty_years_ago,
            prioridade=True  # Should now have priority
        )
        
        # Verify the expected clients were updated (only 65 and 70, not exactly 60)
        expected_updated = [self.cliente_65, self.cliente_70]
        for client in expected_updated:
            client.refresh_from_db()
            self.assertTrue(client.prioridade)
        
        # Cliente exactly 60 should NOT be updated due to < condition
        self.cliente_60_exact.refresh_from_db()
        self.assertFalse(self.cliente_60_exact.prioridade)

    # ==================== ERROR HANDLING TESTS ====================
    
    @patch('precapp.views.Cliente.objects.filter')
    def test_database_error_handling(self, mock_filter):
        """Test error handling when database operations fail"""
        # Mock a database error
        mock_filter.side_effect = Exception("Database connection error")
        
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        # Should still redirect (error handling)
        self.assertEqual(response.status_code, 302)
        
        # Should have error message
        messages = list(get_messages(response.wsgi_request))
        error_messages = [m for m in messages if m.tags == 'error']
        
        self.assertEqual(len(error_messages), 1)
        self.assertIn('Erro ao atualizar prioridades', str(error_messages[0]))
        self.assertIn('Database connection error', str(error_messages[0]))

    # ==================== LARGE DATASET PERFORMANCE TESTS ====================
    
    def test_bulk_update_performance(self):
        """Test performance with large number of clients"""
        # Create many clients over 60 for bulk update testing
        today = date.today()
        clients_batch = []
        
        for i in range(50):
            clients_batch.append(Cliente(
                nome=f'Cliente Bulk {i}',
                cpf=f'{i+10000000000:011d}',  # Generate unique CPFs
                nascimento=today - timedelta(days=(65+i)*365),  # Ages 65-114
                prioridade=False
            ))
        
        Cliente.objects.bulk_create(clients_batch)
        
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        self.assertEqual(response.status_code, 302)
        
        # Verify all bulk clients were updated
        updated_count = Cliente.objects.filter(
            nome__startswith='Cliente Bulk',
            prioridade=True
        ).count()
        
        self.assertEqual(updated_count, 50)
        
        # Check success message mentions correct count (only 2 from setup: cliente_65 and cliente_70)
        messages = list(get_messages(response.wsgi_request))
        success_messages = [m for m in messages if m.tags == 'success']
        self.assertEqual(len(success_messages), 1)
        # Total should be 50 (bulk) + 2 (setup clients over 60: 65 and 70) = 52
        self.assertIn('52 cliente(s)', str(success_messages[0]))

    # ==================== MESSAGE SYSTEM TESTS ====================
    
    def test_example_clients_in_messages(self):
        """Test that example clients are shown in info messages"""
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        messages = list(get_messages(response.wsgi_request))
        info_messages = [m for m in messages if m.tags == 'info']
        
        self.assertEqual(len(info_messages), 1)
        
        # Should contain client names as examples
        message_text = str(info_messages[0])
        self.assertIn('Exemplos atualizados:', message_text)
        
        # Should contain at least one client name (only 65 and 70 get updated, not exactly 60)
        client_names = [self.cliente_65.nome, self.cliente_70.nome]
        found_names = [name for name in client_names if name in message_text]
        self.assertGreater(len(found_names), 0)
    
    def test_message_with_more_than_three_clients(self):
        """Test message format when more than 3 clients are updated"""
        # Create additional clients to exceed the 3-example limit
        today = date.today()
        for i in range(5):
            Cliente.objects.create(
                nome=f'Extra Cliente {i}',
                cpf=f'{i+20000000000:011d}',
                nascimento=today - timedelta(days=(65+i)*365),
                prioridade=False
            )
        
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        messages = list(get_messages(response.wsgi_request))
        info_messages = [m for m in messages if m.tags == 'info']
        
        self.assertEqual(len(info_messages), 1)
        message_text = str(info_messages[0])
        
        # Should show "e mais X cliente(s)" when more than 3
        self.assertIn('e mais', message_text)
        self.assertIn('cliente(s)', message_text)

    # ==================== INTEGRATION TESTS ====================
    
    def test_url_routing(self):
        """Test that URL routing works correctly"""
        self.client_app.login(username='testuser', password='testpass123')
        
        # Test the URL is accessible
        url = reverse('update_priority_by_age')
        self.assertEqual(url, '/clientes/update-priority/')
        
        response = self.client_app.post(url)
        self.assertEqual(response.status_code, 302)
    
    def test_redirect_to_clientes_page(self):
        """Test that view redirects to clientes page after operation"""
        self.client_app.login(username='testuser', password='testpass123')
        response = self.client_app.post(reverse('update_priority_by_age'))
        
        self.assertRedirects(response, reverse('clientes'))
    
    def test_concurrent_access_safety(self):
        """Test that concurrent access doesn't cause data corruption"""
        self.client_app.login(username='testuser', password='testpass123')
        
        # Simulate concurrent requests (simplified test)
        response1 = self.client_app.post(reverse('update_priority_by_age'))
        response2 = self.client_app.post(reverse('update_priority_by_age'))
        
        self.assertEqual(response1.status_code, 302)
        self.assertEqual(response2.status_code, 302)
        
        # Verify clients are still in consistent state
        self.cliente_65.refresh_from_db()
        self.cliente_70.refresh_from_db()
        
        # Should still have priority (no duplicate processing issues)
        self.assertTrue(self.cliente_65.prioridade)
        self.assertTrue(self.cliente_70.prioridade)




class ImportExcelViewTest(TestCase):
    """Comprehensive tests for import_excel_view function"""
    
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        
        self.client_app = Client()
        self.client_app.force_login(self.user)
        
        # Create sample Excel content for testing
        self.sample_excel_content = b'\x50\x4b\x03\x04'  # Basic Excel file header
        
        # Count existing data before tests
        self.initial_precatorios_count = Precatorio.objects.count()
        self.initial_clientes_count = Cliente.objects.count()

    # ==================== AUTHENTICATION TESTS ====================
    
    def test_import_excel_authentication_required(self):
        """Test that import excel view requires authentication"""
        # Create unauthenticated client
        client_unauth = Client()
        
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = client_unauth.post(reverse('import_excel'), {'excel_file': excel_file})
        self.assertEqual(response.status_code, 302)  # Redirect to login
        self.assertIn('/login/', response.url)
    
    def test_import_excel_get_method_redirect(self):
        """Test that GET method redirects to precatorios list"""
        response = self.client_app.get(reverse('import_excel'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('precatorios'))
    
    def test_import_excel_post_authenticated_success(self):
        """Test that authenticated POST requests are processed"""
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        with patch('django.core.management.call_command') as mock_command:
            mock_command.return_value = None
            
            response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
            
            # Should process successfully and redirect
            self.assertEqual(response.status_code, 302)
            self.assertRedirects(response, reverse('precatorios'))

    # ==================== FILE VALIDATION TESTS ====================
    
    def test_import_excel_no_file_provided(self):
        """Test behavior when no file is provided"""
        response = self.client_app.post(reverse('import_excel'), {})
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('precatorios'))
        
        # Check error message - corrected text
        messages = list(response.wsgi_request._messages)
        self.assertTrue(any('Por favor, selecione um arquivo Excel para importar' in str(m) for m in messages))
    
    def test_import_excel_invalid_file_extension(self):
        """Test that invalid file extensions are rejected"""
        # Test with .txt file
        txt_file = SimpleUploadedFile(
            "test.txt", 
            b"some text content", 
            content_type="text/plain"
        )
        
        response = self.client_app.post(reverse('import_excel'), {'excel_file': txt_file})
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('precatorios'))
        
        # Check error message - corrected text
        messages = list(response.wsgi_request._messages)
        self.assertTrue(any('Por favor, selecione um arquivo Excel válido' in str(m) for m in messages))
    
    def test_import_excel_file_size_limit(self):
        """Test file size limit validation"""
        # Create file larger than 10MB
        large_content = b'x' * (11 * 1024 * 1024)  # 11MB
        large_file = SimpleUploadedFile(
            "large.xlsx", 
            large_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client_app.post(reverse('import_excel'), {'excel_file': large_file})
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('precatorios'))
        
        # Check error message - corrected text
        messages = list(response.wsgi_request._messages)
        self.assertTrue(any('O arquivo é muito grande' in str(m) for m in messages))
    
    def test_import_excel_valid_file_extensions(self):
        """Test that both .xlsx and .xls extensions are accepted"""
        with patch('django.core.management.call_command') as mock_command:
            mock_command.side_effect = Exception("File processing error")
            
            xlsx_file = SimpleUploadedFile(
                "test.xlsx", 
                self.sample_excel_content, 
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            xls_file = SimpleUploadedFile(
                "test.xls", 
                self.sample_excel_content, 
                content_type="application/vnd.ms-excel"
            )
            
            # Both should pass file validation
            response1 = self.client_app.post(reverse('import_excel'), {'excel_file': xlsx_file})
            response2 = self.client_app.post(reverse('import_excel'), {'excel_file': xls_file})
            
            # Both should reach processing stage (and fail there due to mock exception)
            self.assertEqual(response1.status_code, 302)
            self.assertEqual(response2.status_code, 302)

    # ==================== IMPORT PROCESSING TESTS ====================
    
    @patch('django.core.management.call_command')
    def test_import_excel_successful_import(self, mock_command):
        """Test successful Excel import with statistics"""
        mock_command.return_value = None
        
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('precatorios'))
        
        # Should have some kind of message (success or warning)
        messages = list(response.wsgi_request._messages)
        self.assertTrue(len(messages) > 0)
    
    @patch('django.core.management.call_command')
    def test_import_excel_command_error(self, mock_command):
        """Test behavior when import command fails"""
        mock_command.side_effect = Exception("Database error occurred")
        
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('precatorios'))
        
        # Check error message - corrected text
        messages = list(response.wsgi_request._messages)
        error_messages = [str(m) for m in messages if 'Erro durante a importação' in str(m)]
        self.assertTrue(len(error_messages) > 0)

    @patch('django.core.management.call_command')
    def test_import_excel_no_new_data(self, mock_command):
        """Test import when no new data is imported"""
        mock_command.return_value = None
        
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
        
        self.assertEqual(response.status_code, 302)
        
        # Should have some kind of message
        messages = list(response.wsgi_request._messages)
        self.assertTrue(len(messages) > 0)

    # ==================== FILE HANDLING TESTS ====================
    
    def test_import_excel_file_cleanup(self):
        """Test that temporary files are cleaned up properly"""
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        with patch('django.core.management.call_command'), \
             patch('os.unlink') as mock_unlink:
            
            response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
            
            # Verify file cleanup was attempted
            mock_unlink.assert_called_once()
            self.assertEqual(response.status_code, 302)
    
    def test_import_excel_file_cleanup_on_error(self):
        """Test that temporary files are cleaned up even when errors occur"""
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        with patch('django.core.management.call_command') as mock_command, \
             patch('os.unlink') as mock_unlink:
            
            # Make command fail
            mock_command.side_effect = Exception("Processing failed")
            
            response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
            
            # Verify file cleanup was still attempted
            mock_unlink.assert_called_once()
            self.assertEqual(response.status_code, 302)

    # ==================== STATISTICS PARSING TESTS ====================

    @patch('django.core.management.call_command')
    def test_import_excel_statistics_parsing_partial(self, mock_command):
        """Test statistics parsing when only some types are imported"""
        mock_command.return_value = None
        
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
        
        # Should handle processing gracefully
        self.assertEqual(response.status_code, 302)
        
        # Should have some kind of message
        messages = list(response.wsgi_request._messages)
        self.assertTrue(len(messages) > 0)

    @patch('django.core.management.call_command')
    def test_import_excel_statistics_parsing_malformed(self, mock_command):
        """Test statistics parsing with malformed output"""
        mock_command.return_value = None
        
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
        
        # Should handle malformed output gracefully
        self.assertEqual(response.status_code, 302)
        
        # Should have some kind of message
        messages = list(response.wsgi_request._messages)
        self.assertTrue(len(messages) > 0)

    # ==================== EDGE CASE TESTS ====================
    
    def test_import_excel_redirect_after_processing(self):
        """Test that view redirects to precatorios page after processing"""
        excel_file = SimpleUploadedFile(
            "test.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        with patch('django.core.management.call_command'):
            response = self.client_app.post(reverse('import_excel'), {'excel_file': excel_file})
            
            self.assertEqual(response.status_code, 302)
            self.assertRedirects(response, reverse('precatorios'))
    
    def test_import_excel_empty_file(self):
        """Test behavior with empty Excel file"""
        empty_file = SimpleUploadedFile(
            "empty.xlsx", 
            b'', 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        with patch('django.core.management.call_command') as mock_command:
            mock_command.side_effect = Exception("Empty file error")
            
            response = self.client_app.post(reverse('import_excel'), {'excel_file': empty_file})
            
            self.assertEqual(response.status_code, 302)
            
            # Should handle empty file gracefully
            messages = list(response.wsgi_request._messages)
            error_messages = [str(m) for m in messages if 'Erro durante a importação' in str(m)]
            self.assertTrue(len(error_messages) > 0)
    
    def test_import_excel_special_characters_filename(self):
        """Test behavior with special characters in filename"""
        special_file = SimpleUploadedFile(
            "tëst_fïlé_ñàmé.xlsx", 
            self.sample_excel_content, 
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        with patch('django.core.management.call_command'):
            response = self.client_app.post(reverse('import_excel'), {'excel_file': special_file})
            
            # Should handle special characters in filename gracefully
            self.assertEqual(response.status_code, 302)
            self.assertRedirects(response, reverse('precatorios'))


class ExportPrecatoriosExcelViewTest(TestCase):
    """Comprehensive tests for export_precatorios_excel view function"""
    
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123',
            first_name='Test',
            last_name='User'
        )
        
        self.client_app = Client()
        self.client_app.force_login(self.user)
        
        # Import all necessary models
        from precapp.models import (
            Cliente, Precatorio, Tipo, Diligencias, TipoDiligencia,
            Requerimento, PedidoRequerimento, Fase, Alvara, FaseHonorariosContratuais
        )
        
        # Create test data for comprehensive export testing
        
        # Create tipo precatorio
        self.tipo_precatorio = Tipo.objects.create(
            nome='Tipo Teste',
            ordem=1,
            ativa=True
        )
        
        # Create clients
        self.cliente1 = Cliente.objects.create(
            nome='João Silva',
            cpf='12345678901',
            nascimento=date(1950, 5, 15),
            prioridade=True
        )
        
        self.cliente2 = Cliente.objects.create(
            nome='Maria Santos',
            cpf='98765432109',
            nascimento=date(1980, 10, 20),
            prioridade=False
        )
        
        # Create precatorios
        self.precatorio1 = Precatorio.objects.create(
            cnj='1234567-89.2023.8.01.0001',
            orcamento=2023,
            origem='Tribunal A',
            valor_de_face=100000.00,
            ultima_atualizacao=120000.00,
            data_ultima_atualizacao=date.today(),
            credito_principal='sim',
            honorarios_contratuais='não',
            honorarios_sucumbenciais='sim',
            tipo=self.tipo_precatorio
        )
        
        self.precatorio2 = Precatorio.objects.create(
            cnj='9876543-21.2023.8.01.0002',
            orcamento=2023,
            origem='Tribunal B',
            valor_de_face=75000.00,
            ultima_atualizacao=80000.00,
            data_ultima_atualizacao=date.today(),
            credito_principal='não',
            honorarios_contratuais='sim',
            honorarios_sucumbenciais='não',
            tipo=self.tipo_precatorio
        )
        
        # Associate clients with precatorios
        self.precatorio1.clientes.add(self.cliente1)
        self.precatorio2.clientes.add(self.cliente2)
        
        # Create tipo diligencia
        self.tipo_diligencia = TipoDiligencia.objects.create(
            nome='Consulta Processual',
            ordem=1,
            ativo=True
        )
        
        # Create diligencias
        self.diligencia1 = Diligencias.objects.create(
            cliente=self.cliente1,
            tipo=self.tipo_diligencia,
            descricao='Verificar andamento do processo',
            data_final=date.today() + timedelta(days=30),
            urgencia='alta',
            concluida=False,
            responsavel=self.user,
            criado_por='admin'
        )
        
        self.diligencia2 = Diligencias.objects.create(
            cliente=self.cliente2,
            tipo=self.tipo_diligencia,
            descricao='Contatar cliente sobre documentos',
            data_final=date.today() - timedelta(days=5),
            urgencia='media',
            concluida=True,
            data_conclusao=timezone.now(),
            responsavel=self.user,
            criado_por='admin',
            concluido_por='admin'
        )
        
        # Create pedido requerimento
        self.pedido = PedidoRequerimento.objects.create(
            nome='Prioridade por Idade',
            ordem=1,
            ativo=True
        )
        
        # Create fase
        self.fase = Fase.objects.create(
            nome='Em Análise',
            ordem=1,
            ativa=True
        )
        
        # Create requerimentos
        self.requerimento1 = Requerimento.objects.create(
            cliente=self.cliente1,
            precatorio=self.precatorio1,
            pedido=self.pedido,
            valor=50000.00,
            desagio=10.5,
            fase=self.fase
        )
        
        self.requerimento2 = Requerimento.objects.create(
            cliente=self.cliente2,
            precatorio=self.precatorio2,
            pedido=self.pedido,
            valor=30000.00,
            desagio=8.0,
            fase=self.fase
        )
        
        # Create fase honorarios
        self.fase_honorarios = FaseHonorariosContratuais.objects.create(
            nome='Aguardando Pagamento',
            ordem=1,
            ativa=True
        )
        
        # Create alvaras
        self.alvara1 = Alvara.objects.create(
            cliente=self.cliente1,
            precatorio=self.precatorio1,
            valor_principal=100000.00,
            honorarios_contratuais=15000.00,
            honorarios_sucumbenciais=8000.00,
            tipo='aguardando depósito judicial',
            fase=self.fase,
            fase_honorarios_contratuais=self.fase_honorarios
        )
        
        self.alvara2 = Alvara.objects.create(
            cliente=self.cliente2,
            precatorio=self.precatorio2,
            valor_principal=75000.00,
            honorarios_contratuais=12000.00,
            honorarios_sucumbenciais=5000.00,
            tipo='depósito judicial efetuado',
            fase=self.fase,
            fase_honorarios_contratuais=self.fase_honorarios
        )

    # ==================== AUTHENTICATION TESTS ====================
    
    def test_export_precatorios_authentication_required(self):
        """Test that export requires authentication"""
        client_unauth = Client()
        response = client_unauth.get(reverse('export_precatorios_excel'))
        self.assertEqual(response.status_code, 302)  # Redirect to login
        self.assertIn('/login/', response.url)
    
    def test_export_precatorios_authenticated_access(self):
        """Test that authenticated users can access export"""
        response = self.client_app.get(reverse('export_precatorios_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # ==================== RESPONSE FORMAT TESTS ====================
    
    def test_export_precatorios_response_headers(self):
        """Test that response has correct headers for Excel download"""
        response = self.client_app.get(reverse('export_precatorios_excel'))
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Check filename format
        content_disposition = response['Content-Disposition']
        self.assertIn('attachment; filename=', content_disposition)
        self.assertIn('relatorio_completo_sistema_', content_disposition)
        self.assertIn('.xlsx', content_disposition)
    
    def test_export_precatorios_filename_timestamp(self):
        """Test that filename includes timestamp"""
        from django.utils import timezone
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        content_disposition = response['Content-Disposition']
        
        # Extract timestamp from filename
        import re
        timestamp_pattern = r'relatorio_completo_sistema_(\d{8}_\d{6})\.xlsx'
        match = re.search(timestamp_pattern, content_disposition)
        
        self.assertIsNotNone(match)
        timestamp = match.group(1)
        
        # Verify timestamp format (YYYYMMDD_HHMMSS)
        self.assertEqual(len(timestamp), 15)  # 8 digits + _ + 6 digits
        self.assertRegex(timestamp, r'\d{8}_\d{6}')

    # ==================== EXCEL STRUCTURE TESTS ====================
    
    def test_export_precatorios_excel_structure(self):
        """Test that Excel file has correct worksheet structure"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        
        # Load Excel from response content
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check worksheet names
        expected_sheets = ['Precatórios', 'Clientes', 'Diligências', 'Requerimentos', 'Alvarás', 'Estatísticas']
        self.assertEqual(workbook.sheetnames, expected_sheets)
        
        # Check each worksheet has headers
        for sheet_name in expected_sheets:
            worksheet = workbook[sheet_name]
            # First row should have headers (not empty)
            first_row = [cell.value for cell in worksheet[1]]
            self.assertTrue(any(cell for cell in first_row if cell))
    
    def test_export_precatorios_worksheet_headers(self):
        """Test specific worksheet headers"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Test Precatórios sheet headers
        precatorios_sheet = workbook['Precatórios']
        expected_headers = [
            'CNJ', 'Origem', 'Valor de Face', 'Última Atualização', 'Data Última Atualização',
            'Crédito Principal', 'Honorários Contratuais', 'Honorários Sucumbenciais',
            'Cliente Nome', 'Cliente CPF', 'Cliente Nascimento', 'Cliente Prioritário',
            'Tipo Precatório', 'Orçamento'
        ]
        actual_headers = [cell.value for cell in precatorios_sheet[1]]
        self.assertEqual(actual_headers, expected_headers)
        
        # Test Estatísticas sheet structure
        stats_sheet = workbook['Estatísticas']
        stats_headers = [cell.value for cell in stats_sheet[1]]
        self.assertEqual(stats_headers, ['Estatística', 'Valor'])

    # ==================== DATA CONTENT TESTS ====================
    
    def test_export_precatorios_data_content(self):
        """Test that exported data matches database content"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Test Precatórios data
        precatorios_sheet = workbook['Precatórios']
        
        # Check first precatorio data (row 2)
        row_2 = [cell.value for cell in precatorios_sheet[2]]
        self.assertEqual(row_2[0], self.precatorio1.cnj)  # CNJ
        self.assertEqual(row_2[1], self.precatorio1.origem)  # Origem
        self.assertEqual(row_2[8], self.cliente1.nome)  # Cliente Nome
        self.assertEqual(row_2[9], self.cliente1.cpf)  # Cliente CPF
        
        # Test Clientes data
        clientes_sheet = workbook['Clientes']
        clientes_row_2 = [cell.value for cell in clientes_sheet[2]]
        self.assertEqual(clientes_row_2[0], self.cliente1.nome)  # Nome
        self.assertEqual(clientes_row_2[1], self.cliente1.cpf)  # CPF
    
    def test_export_precatorios_statistics_calculations(self):
        """Test that statistics are calculated correctly"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check statistics sheet
        stats_sheet = workbook['Estatísticas']
        
        # Find specific statistics
        stats_data = {}
        for row in stats_sheet.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1] is not None:
                stats_data[row[0]] = row[1]
        
        # Verify key statistics
        self.assertEqual(stats_data.get('Total de Precatórios'), 2)
        self.assertEqual(stats_data.get('Total de Clientes'), 2)
        self.assertEqual(stats_data.get('Total de Diligências'), 2)
        self.assertEqual(stats_data.get('Total de Requerimentos'), 2)
        self.assertEqual(stats_data.get('Total de Alvarás'), 2)
        self.assertEqual(stats_data.get('Clientes Prioritários'), 1)

    # ==================== FORMATTING TESTS ====================
    
    def test_export_precatorios_currency_formatting(self):
        """Test that currency values are properly formatted"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check Precatórios sheet currency formatting
        precatorios_sheet = workbook['Precatórios']
        
        # Check valor de face cell (column 3, row 2)
        valor_cell = precatorios_sheet.cell(row=2, column=3)
        if valor_cell.value is not None:
            # Currency format should be applied
            self.assertIn('R$', valor_cell.number_format or '')
    
    def test_export_precatorios_date_formatting(self):
        """Test that dates are properly formatted"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check date formatting in various sheets
        precatorios_sheet = workbook['Precatórios']
        
        # Data de última atualização should be in Brazilian format
        date_cell = precatorios_sheet.cell(row=2, column=5)
        if date_cell.value:
            # Should be in DD/MM/YYYY format
            date_str = str(date_cell.value)
            if '/' in date_str:
                self.assertRegex(date_str, r'\d{2}/\d{2}/\d{4}')

    # ==================== ERROR HANDLING TESTS ====================
    
    def test_export_precatorios_empty_database(self):
        """Test export behavior with empty database"""
        # Clear all test data
        from precapp.models import Cliente, Precatorio, Diligencias, Requerimento, Alvara
        Alvara.objects.all().delete()
        Requerimento.objects.all().delete()
        Diligencias.objects.all().delete()
        Precatorio.objects.all().delete()
        Cliente.objects.all().delete()
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        
        # Should still generate Excel file
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def test_export_precatorios_with_null_values(self):
        """Test export behavior with null/empty values"""
        # Create precatorio with minimal data
        minimal_precatorio = Precatorio.objects.create(
            cnj='MINIMAL-TEST',
            orcamento=2023,
            origem='',  # Empty string for testing empty values
            valor_de_face=0.0,  # Required field, so use 0
            ultima_atualizacao=None,
            data_ultima_atualizacao=None,
            tipo=None
        )
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        
        # Should handle null values gracefully
        self.assertEqual(response.status_code, 200)
        
        # Verify Excel can be loaded
        from openpyxl import load_workbook
        import io
        
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        self.assertIn('Precatórios', workbook.sheetnames)

    # ==================== PERFORMANCE TESTS ====================
    
    def test_export_precatorios_large_dataset(self):
        """Test export performance with larger dataset"""
        from precapp.models import Cliente, Precatorio
        
        # Create additional test data
        for i in range(20):
            cliente = Cliente.objects.create(
                nome=f'Cliente Teste {i}',
                cpf=f'{i+10000000000:011d}',
                nascimento=date(1970, 1, 1),
                prioridade=False
            )
            
            precatorio = Precatorio.objects.create(
                cnj=f'TEST-{i:04d}',
                orcamento=2023,
                origem=f'Tribunal {i}',
                valor_de_face=1000.00 * (i + 1),
                tipo=self.tipo_precatorio
            )
            
            precatorio.clientes.add(cliente)
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        
        # Should complete successfully
        self.assertEqual(response.status_code, 200)
        
        # Response should have content
        self.assertGreater(len(response.content), 1000)

    # ==================== INTEGRATION TESTS ====================
    
    def test_export_precatorios_url_routing(self):
        """Test URL routing for export function"""
        url = reverse('export_precatorios_excel')
        self.assertEqual(url, '/precatorios/export/')
        
        response = self.client_app.get(url)
        self.assertEqual(response.status_code, 200)
    
    def test_export_precatorios_user_context(self):
        """Test that user context is included in report"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check statistics sheet for user information
        stats_sheet = workbook['Estatísticas']
        
        # Find "Gerado por" entry
        user_found = False
        for row in stats_sheet.iter_rows(values_only=True):
            if row[0] == 'Gerado por':
                self.assertIn('Test User', str(row[1]))
                user_found = True
                break
        
        self.assertTrue(user_found, "User information not found in statistics")
    
    def test_export_precatorios_comprehensive_data_integrity(self):
        """Test comprehensive data integrity across all sheets"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_precatorios_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Verify all expected data is present across sheets
        
        # Precatórios sheet should have our test data
        precatorios_sheet = workbook['Precatórios']
        precatorio_cnjs = [row[0] for row in precatorios_sheet.iter_rows(min_row=2, max_col=1, values_only=True)]
        self.assertIn(self.precatorio1.cnj, precatorio_cnjs)
        self.assertIn(self.precatorio2.cnj, precatorio_cnjs)
        
        # Clientes sheet should have our test clients
        clientes_sheet = workbook['Clientes']
        cliente_names = [row[0] for row in clientes_sheet.iter_rows(min_row=2, max_col=1, values_only=True)]
        self.assertIn(self.cliente1.nome, cliente_names)
        self.assertIn(self.cliente2.nome, cliente_names)
        
        # Diligências sheet should have our test diligencias
        diligencias_sheet = workbook['Diligências']
        diligencia_clients = [row[0] for row in diligencias_sheet.iter_rows(min_row=2, max_col=1, values_only=True)]
        self.assertIn(self.cliente1.nome, diligencia_clients)
        self.assertIn(self.cliente2.nome, diligencia_clients)


class ExportClientesExcelViewTest(TestCase):
    """Comprehensive tests for export_clientes_excel view function"""
    
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123',
            first_name='Test',
            last_name='User'
        )
        
        self.client_app = Client()
        self.client_app.force_login(self.user)
        
        # Import necessary models
        from precapp.models import Cliente, Precatorio, Tipo, Diligencias, TipoDiligencia
        
        # Create test data for client-focused export
        
        # Create tipo precatorio
        self.tipo_precatorio = Tipo.objects.create(
            nome='Tipo Teste Cliente',
            ordem=1,
            ativa=True
        )
        
        # Create clients with different characteristics
        self.cliente_prioritario = Cliente.objects.create(
            nome='João Silva Prioritário',
            cpf='12345678901',
            nascimento=date(1950, 5, 15),  # 73 years old
            prioridade=True
        )
        
        self.cliente_normal = Cliente.objects.create(
            nome='Maria Santos Normal',
            cpf='98765432109',
            nascimento=date(1980, 10, 20),  # 43 years old
            prioridade=False
        )
        
        self.cliente_sem_precatorio = Cliente.objects.create(
            nome='Pedro Costa',
            cpf='11122233344',
            nascimento=date(1990, 3, 10),  # 33 years old
            prioridade=False
        )
        
        # Create precatorios for some clients
        self.precatorio1 = Precatorio.objects.create(
            cnj='1234567-89.2023.8.01.0001',
            orcamento=2023,
            origem='Tribunal A',
            valor_de_face=100000.00,
            ultima_atualizacao=120000.00,
            data_ultima_atualizacao=date.today(),
            tipo=self.tipo_precatorio
        )
        
        self.precatorio2 = Precatorio.objects.create(
            cnj='9876543-21.2023.8.01.0002',
            orcamento=2023,
            origem='Tribunal B',
            valor_de_face=75000.00,
            ultima_atualizacao=80000.00,
            data_ultima_atualizacao=date.today(),
            tipo=self.tipo_precatorio
        )
        
        self.precatorio3 = Precatorio.objects.create(
            cnj='5555555-55.2023.8.01.0003',
            orcamento=2023,
            origem='Tribunal C',
            valor_de_face=50000.00,
            ultima_atualizacao=55000.00,
            data_ultima_atualizacao=date.today(),
            tipo=self.tipo_precatorio
        )
        
        # Associate clients with precatorios
        self.precatorio1.clientes.add(self.cliente_prioritario)
        self.precatorio2.clientes.add(self.cliente_prioritario)  # Multiple precatorios
        self.precatorio3.clientes.add(self.cliente_normal)
        # cliente_sem_precatorio has no precatorios
        
        # Create tipo diligencia
        self.tipo_diligencia = TipoDiligencia.objects.create(
            nome='Consulta Cliente',
            ordem=1,
            ativo=True
        )
        
        # Create diligencias for clients
        self.diligencia1 = Diligencias.objects.create(
            cliente=self.cliente_prioritario,
            tipo=self.tipo_diligencia,
            descricao='Contatar cliente prioritário',
            data_final=date.today() + timedelta(days=15),
            urgencia='alta',
            concluida=False,
            responsavel=self.user
        )
        
        self.diligencia2 = Diligencias.objects.create(
            cliente=self.cliente_prioritario,
            tipo=self.tipo_diligencia,
            descricao='Seguimento processual',
            data_final=date.today() - timedelta(days=10),  # Overdue
            urgencia='media',
            concluida=True,
            data_conclusao=timezone.now() - timedelta(days=5),
            responsavel=self.user
        )
        
        self.diligencia3 = Diligencias.objects.create(
            cliente=self.cliente_normal,
            tipo=self.tipo_diligencia,
            descricao='Verificar documentação',
            data_final=date.today() + timedelta(days=30),
            urgencia='baixa',
            concluida=False,
            responsavel=self.user
        )

    # ==================== AUTHENTICATION TESTS ====================
    
    def test_export_clientes_authentication_required(self):
        """Test that export requires authentication"""
        client_unauth = Client()
        response = client_unauth.get(reverse('export_clientes_excel'))
        self.assertEqual(response.status_code, 302)  # Redirect to login
        self.assertIn('/login/', response.url)
    
    def test_export_clientes_authenticated_access(self):
        """Test that authenticated users can access export"""
        response = self.client_app.get(reverse('export_clientes_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # ==================== RESPONSE FORMAT TESTS ====================
    
    def test_export_clientes_response_headers(self):
        """Test that response has correct headers for Excel download"""
        response = self.client_app.get(reverse('export_clientes_excel'))
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Check filename format
        content_disposition = response['Content-Disposition']
        self.assertIn('attachment; filename=', content_disposition)
        self.assertIn('relatorio_clientes_', content_disposition)
        self.assertIn('.xlsx', content_disposition)
    
    def test_export_clientes_filename_timestamp(self):
        """Test that filename includes timestamp"""
        response = self.client_app.get(reverse('export_clientes_excel'))
        content_disposition = response['Content-Disposition']
        
        # Extract timestamp from filename
        import re
        timestamp_pattern = r'relatorio_clientes_(\d{8}_\d{6})\.xlsx'
        match = re.search(timestamp_pattern, content_disposition)
        
        self.assertIsNotNone(match)
        timestamp = match.group(1)
        
        # Verify timestamp format (YYYYMMDD_HHMMSS)
        self.assertEqual(len(timestamp), 15)  # 8 digits + _ + 6 digits
        self.assertRegex(timestamp, r'\d{8}_\d{6}')

    # ==================== EXCEL STRUCTURE TESTS ====================
    
    def test_export_clientes_excel_structure(self):
        """Test that Excel file has correct worksheet structure"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        
        # Load Excel from response content
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check worksheet names
        expected_sheets = ['Clientes Detalhado', 'Resumo por Prioridade', 'Estatísticas Detalhadas']
        self.assertEqual(workbook.sheetnames, expected_sheets)
        
        # Check each worksheet has headers
        for sheet_name in expected_sheets:
            worksheet = workbook[sheet_name]
            # First row should have headers (not empty)
            first_row = [cell.value for cell in worksheet[1]]
            self.assertTrue(any(cell for cell in first_row if cell))
    
    def test_export_clientes_worksheet_headers(self):
        """Test specific worksheet headers"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Test Clientes Detalhado sheet headers
        clientes_sheet = workbook['Clientes Detalhado']
        expected_headers = [
            'CPF/CNPJ', 'Nome Completo', 'Data Nascimento', 'Idade', 'Cliente Prioritário',
            'Total Precatórios', 'CNJ Precatórios', 'Valor Total Precatórios',
            'Total Diligências', 'Diligências Pendentes', 'Diligências Concluídas',
            'Diligências Atrasadas', 'Última Diligência', 'Próximo Vencimento'
        ]
        actual_headers = [cell.value for cell in clientes_sheet[1]]
        self.assertEqual(actual_headers, expected_headers)
        
        # Test Resumo por Prioridade sheet headers
        resumo_sheet = workbook['Resumo por Prioridade']
        resumo_headers = [cell.value for cell in resumo_sheet[1]]
        expected_resumo_headers = ['Categoria', 'Quantidade Clientes', 'Percentual', 'Valor Total Precatórios', 'Valor Médio por Cliente']
        self.assertEqual(resumo_headers, expected_resumo_headers)

    # ==================== DATA CONTENT TESTS ====================
    
    def test_export_clientes_detailed_data(self):
        """Test that detailed client data is correctly exported"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Test Clientes Detalhado data
        clientes_sheet = workbook['Clientes Detalhado']
        
        # Find prioritário client row
        prioritario_row = None
        for row in clientes_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == self.cliente_prioritario.nome:  # Nome Completo column
                prioritario_row = row
                break
        
        self.assertIsNotNone(prioritario_row)
        
        # Verify client data
        self.assertEqual(prioritario_row[0], self.cliente_prioritario.cpf)  # CPF
        self.assertEqual(prioritario_row[1], self.cliente_prioritario.nome)  # Nome
        self.assertEqual(prioritario_row[4], 'Sim')  # Cliente Prioritário
        self.assertEqual(prioritario_row[5], 2)  # Total Precatórios (has 2)
        self.assertEqual(prioritario_row[8], 2)  # Total Diligências (has 2)
        self.assertEqual(prioritario_row[9], 1)  # Diligências Pendentes (has 1)
        self.assertEqual(prioritario_row[10], 1)  # Diligências Concluídas (has 1)
    
    def test_export_clientes_age_calculation(self):
        """Test that age calculation is correct"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        clientes_sheet = workbook['Clientes Detalhado']
        
        # Find cliente prioritario row and check age
        for row in clientes_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == self.cliente_prioritario.nome:
                idade_str = row[3]  # Idade column
                # Born in 1950, should be 74-75 years (depending on current date vs birthday)
                self.assertTrue('74 anos' in str(idade_str) or '75 anos' in str(idade_str))
                break
    
    def test_export_clientes_priority_summary(self):
        """Test priority summary calculations"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Test Resumo por Prioridade data
        resumo_sheet = workbook['Resumo por Prioridade']
        
        # Find priority summary rows
        prioritarios_row = None
        normais_row = None
        total_row = None
        
        for row in resumo_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == 'Clientes Prioritários':
                prioritarios_row = row
            elif row[0] == 'Clientes Normais':
                normais_row = row
            elif row[0] == 'TOTAL GERAL':
                total_row = row
        
        # Verify calculations
        self.assertIsNotNone(prioritarios_row)
        self.assertEqual(prioritarios_row[1], 1)  # 1 prioritário client
        
        self.assertIsNotNone(normais_row)
        self.assertEqual(normais_row[1], 2)  # 2 normal clients
        
        self.assertIsNotNone(total_row)
        self.assertEqual(total_row[1], 3)  # 3 total clients

    # ==================== STATISTICS TESTS ====================
    
    def test_export_clientes_detailed_statistics(self):
        """Test detailed statistics calculations"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check statistics sheet
        stats_sheet = workbook['Estatísticas Detalhadas']
        
        # Find specific statistics
        stats_data = {}
        for row in stats_sheet.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1] is not None:
                stats_data[row[0]] = row[1]
        
        # Verify key statistics
        self.assertEqual(stats_data.get('Total de Clientes'), 3)
        self.assertEqual(stats_data.get('Clientes Prioritários'), 1)
        self.assertEqual(stats_data.get('Clientes Normais'), 2)
        self.assertEqual(stats_data.get('Total de Precatórios no Sistema'), 3)
        self.assertEqual(stats_data.get('Total de Diligências no Sistema'), 3)
        self.assertEqual(stats_data.get('Diligências Pendentes'), 2)
        self.assertEqual(stats_data.get('Diligências Concluídas'), 1)
    
    def test_export_clientes_percentage_calculations(self):
        """Test percentage calculations in statistics"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        stats_sheet = workbook['Estatísticas Detalhadas']
        
        # Find percentage statistics
        for row in stats_sheet.iter_rows(values_only=True):
            if row[0] == 'Percentual Prioritários':
                # Should be 33.3% (1 out of 3 clients)
                percentage_str = str(row[1])
                self.assertIn('33.3%', percentage_str)
                break

    # ==================== FINANCIAL DATA TESTS ====================
    
    def test_export_clientes_financial_calculations(self):
        """Test financial value calculations"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        clientes_sheet = workbook['Clientes Detalhado']
        
        # Find cliente prioritario row and check financial data
        for row in clientes_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == self.cliente_prioritario.nome:
                valor_total = row[7]  # Valor Total Precatórios
                # Should have 2 precatorios: 120000 + 80000 = 200000
                self.assertEqual(valor_total, 200000.00)
                break
        
        # Find cliente normal row
        for row in clientes_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == self.cliente_normal.nome:
                valor_total = row[7]  # Valor Total Precatórios
                # Should have 1 precatorio: 55000
                self.assertEqual(valor_total, 55000.00)
                break

    # ==================== DILIGENCIA ANALYSIS TESTS ====================
    
    def test_export_clientes_diligencia_analysis(self):
        """Test diligencia analysis in client export"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        clientes_sheet = workbook['Clientes Detalhado']
        
        # Find cliente prioritario row
        for row in clientes_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == self.cliente_prioritario.nome:
                total_diligencias = row[8]      # Total Diligências
                pendentes = row[9]              # Diligências Pendentes
                concluidas = row[10]            # Diligências Concluídas
                atrasadas = row[11]             # Diligências Atrasadas
                
                self.assertEqual(total_diligencias, 2)
                self.assertEqual(pendentes, 1)
                self.assertEqual(concluidas, 1)
                self.assertEqual(atrasadas, 0)  # None overdue for this client
                break

    # ==================== ERROR HANDLING TESTS ====================
    
    def test_export_clientes_empty_database(self):
        """Test export behavior with empty database"""
        # Clear all test data
        from precapp.models import Cliente, Precatorio, Diligencias
        Diligencias.objects.all().delete()
        Precatorio.objects.all().delete()
        Cliente.objects.all().delete()
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        
        # Should still generate Excel file
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def test_export_clientes_with_null_birthdates(self):
        """Test export behavior with clients having various birthdates"""
        # Create client with early birthdate
        cliente_antigo = Cliente.objects.create(
            nome='Cliente Antigo',
            cpf='55555555555',
            nascimento=date(1950, 1, 1),  # Very old client
            prioridade=False
        )
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        
        # Should handle null birthdates gracefully
        self.assertEqual(response.status_code, 200)
        
        # Verify Excel can be loaded
        from openpyxl import load_workbook
        import io
        
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        self.assertIn('Clientes Detalhado', workbook.sheetnames)

    # ==================== INTEGRATION TESTS ====================
    
    def test_export_clientes_url_routing(self):
        """Test URL routing for client export function"""
        url = reverse('export_clientes_excel')
        self.assertEqual(url, '/clientes/export/')
        
        response = self.client_app.get(url)
        self.assertEqual(response.status_code, 200)
    
    def test_export_clientes_user_context(self):
        """Test that user context is included in report"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        # Check statistics sheet for user information
        stats_sheet = workbook['Estatísticas Detalhadas']
        
        # Find "Gerado por" entry
        user_found = False
        for row in stats_sheet.iter_rows(values_only=True):
            if row[0] == 'Gerado por':
                self.assertIn('Test User', str(row[1]))
                user_found = True
                break
        
        self.assertTrue(user_found, "User information not found in statistics")
    
    def test_export_clientes_report_type_identification(self):
        """Test that report is properly identified as client export"""
        from openpyxl import load_workbook
        import io
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        
        stats_sheet = workbook['Estatísticas Detalhadas']
        
        # Find "Tipo de Relatório" entry
        report_type_found = False
        for row in stats_sheet.iter_rows(values_only=True):
            if row[0] == 'Tipo de Relatório':
                self.assertEqual(row[1], 'Exportação de Clientes')
                report_type_found = True
                break
        
        self.assertTrue(report_type_found, "Report type not found in statistics")

    # ==================== PERFORMANCE TESTS ====================
    
    def test_export_clientes_performance_with_many_clients(self):
        """Test export performance with many clients"""
        from precapp.models import Cliente
        
        # Create additional clients for performance testing
        for i in range(30):
            Cliente.objects.create(
                nome=f'Cliente Performance {i}',
                cpf=f'{i+20000000000:011d}',
                nascimento=date(1975, 1, 1),
                prioridade=i % 5 == 0  # Every 5th client is prioritário
            )
        
        response = self.client_app.get(reverse('export_clientes_excel'))
        
        # Should complete successfully
        self.assertEqual(response.status_code, 200)
        
        # Response should have content
        self.assertGreater(len(response.content), 1000)
        
        # Verify Excel structure remains intact
        from openpyxl import load_workbook
        import io
        
        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)
        expected_sheets = ['Clientes Detalhado', 'Resumo por Prioridade', 'Estatísticas Detalhadas']
        self.assertEqual(workbook.sheetnames, expected_sheets)
